"""
Services Blueprint
Handles all consulting services pages and API endpoints
"""

from flask import Blueprint, render_template, session, jsonify, current_app, make_response
from flask_jwt_extended import jwt_required, get_jwt_identity
from models import Service, ServiceOffering, User, Project, AILog
from utils.decorators import login_required
from utils.ai_providers.ai_manager import AIManager
from weasyprint import HTML
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import json
import re

services_bp = Blueprint('services', __name__, url_prefix='/services')

def get_lang():
    """Get current language from session"""
    return session.get('language', 'ar')

def get_db():
    """Get database instance from current app"""
    return current_app.extensions['sqlalchemy']

def get_all_services_with_offerings():
    """Get all active services with their offerings for sidebar"""
    db = get_db()
    services = db.session.query(Service).filter_by(is_active=True).order_by(Service.display_order).all()
    # Use a separate attribute to avoid mutating the SQLAlchemy relationship
    for service in services:
        service.active_offerings = db.session.query(ServiceOffering).filter_by(
            service_id=service.id,
            is_active=True
        ).order_by(ServiceOffering.display_order).all()
    return services

@services_bp.route('/')
def index():
    """Services homepage - shows all categories (public access)"""
    lang = get_lang()
    db = get_db()
    services = db.session.query(Service).filter_by(is_active=True).order_by(Service.display_order).all()
    
    # Get all services for sidebar
    all_services = get_all_services_with_offerings()
    
    return render_template(
        'services/index.html',
        services=services,
        all_services=all_services,
        current_service=None,
        current_offering=None,
        lang=lang
    )

@services_bp.route('/api/all')
def api_get_all_services():
    """API endpoint to get all services with their offerings"""
    lang = get_lang()
    db = get_db()
    services = db.session.query(Service).filter_by(is_active=True).order_by(Service.display_order).all()
    
    result = []
    for service in services:
        service_dict = service.to_dict(lang)
        service_dict['offerings'] = [
            offering.to_dict(lang) 
            for offering in sorted(service.offerings, key=lambda x: x.display_order)
            if offering.is_active
        ]
        result.append(service_dict)
    
    return jsonify(result)

@services_bp.route('/<service_slug>')
def service_detail(service_slug):
    """Service category page - shows all offerings for a service (public access)"""
    lang = get_lang()
    db = get_db()
    service = db.session.query(Service).filter_by(slug=service_slug, is_active=True).first()
    if not service:
        from flask import abort
        abort(404)
    
    # Get active offerings
    offerings = db.session.query(ServiceOffering).filter_by(
        service_id=service.id,
        is_active=True
    ).order_by(ServiceOffering.display_order).all()
    
    # Get all services for sidebar
    all_services = get_all_services_with_offerings()
    
    return render_template(
        'services/service_detail.html',
        service=service,
        offerings=offerings,
        all_services=all_services,
        current_service=service,
        current_offering=None,
        lang=lang
    )

@services_bp.route('/<service_slug>/<offering_slug>')
def offering_detail(service_slug, offering_slug):
    """Service offering page - individual service with AI interaction (public access)"""
    from flask_jwt_extended import verify_jwt_in_request
    
    lang = get_lang()
    db = get_db()
    
    service = db.session.query(Service).filter_by(slug=service_slug, is_active=True).first()
    if not service:
        from flask import abort
        abort(404)
    
    offering = db.session.query(ServiceOffering).filter_by(
        service_id=service.id,
        slug=offering_slug,
        is_active=True
    ).first()
    if not offering:
        from flask import abort
        abort(404)
    
    # Get current user if logged in
    user = None
    projects = []
    try:
        verify_jwt_in_request(optional=True)
        user_id = get_jwt_identity()
        if user_id:
            user = db.session.get(User, int(user_id))
            # Get user's projects for this offering only if logged in
            projects = db.session.query(Project).filter_by(
                user_id=int(user_id),
                module=f"{service_slug}_{offering_slug}"
            ).order_by(Project.updated_at.desc()).limit(5).all()
    except:
        pass
    
    # Get all services for sidebar
    all_services = get_all_services_with_offerings()
    
    return render_template(
        'services/offering_detail.html',
        service=service,
        offering=offering,
        projects=projects,
        all_services=all_services,
        current_service=service,
        current_offering=offering,
        lang=lang
    )

@services_bp.route('/api/<service_slug>/<offering_slug>/generate', methods=['POST'])
@login_required
def api_generate_content(service_slug, offering_slug):
    """API endpoint to generate AI content for an offering"""
    from flask import request, abort
    import json
    
    db = get_db()
    
    # Get offering
    service = db.session.query(Service).filter_by(slug=service_slug).first()
    if not service:
        abort(404)
    
    offering = db.session.query(ServiceOffering).filter_by(
        service_id=service.id,
        slug=offering_slug
    ).first()
    if not offering:
        abort(404)
    
    # Get current user
    user_id = get_jwt_identity()
    user = db.session.get(User, int(user_id))
    
    # Initialize ai_credits_used if None
    if user.ai_credits_used is None:
        user.ai_credits_used = 0
    
    # Check AI credits
    plan = user.plan_ref
    if plan and plan.ai_credits_limit:
        if user.ai_credits_used >= plan.ai_credits_limit:
            return jsonify({'error': 'AI credits limit exceeded'}), 403
    
    # Get form data
    form_data = request.get_json()
    
    # Build prompt from template
    lang = session.get('language', 'ar')
    
    # Use custom prompt template if available, otherwise use default
    if offering.ai_prompt_template:
        # Parse form_fields to extract field names and values
        import html
        form_fields_schema = []
        try:
            if offering.form_fields:
                form_fields_schema = json.loads(offering.form_fields) if isinstance(offering.form_fields, str) else offering.form_fields
        except Exception as e:
            return jsonify({'error': 'Invalid form fields schema'}), 500
        
        # Validate that form_fields_schema is a list
        if not isinstance(form_fields_schema, list):
            form_fields_schema = []
        
        # Build replacement dictionary with field values
        replacement_dict = {
            'project_name': form_data.get('project_name', 'غير محدد / Not specified'),
            'description': form_data.get('description', ''),
            'additional_context': form_data.get('additional_context', '')
        }
        
        # Build structured input context with field metadata
        input_context_parts = []
        
        # Add custom fields to replacement dict with validation
        for field in form_fields_schema:
            if not isinstance(field, dict):
                continue
                
            field_name = field.get('name')
            if not field_name or not isinstance(field_name, str):
                continue
            
            field_type = field.get('type', 'text')
            field_label_ar = field.get('label_ar', field_name)
            field_label_en = field.get('label_en', field_name)
            field_value = form_data.get(field_name, 'N/A')
            field_options = field.get('options', [])
            
            # Validate required fields
            if field.get('required') and (not field_value or field_value == 'N/A'):
                return jsonify({'error': f'Required field missing: {field_name}'}), 400
            
            # Type validation
            if field_type == 'number' and field_value != 'N/A':
                try:
                    field_value = float(field_value)
                except ValueError:
                    return jsonify({'error': f'Invalid number format for field: {field_name}'}), 400
            
            # Sanitize field value to prevent injection
            field_value_str = str(field_value)
            field_value_sanitized = field_value_str.replace('{', '').replace('}', '').strip()
            if len(field_value_sanitized) > 5000:
                field_value_sanitized = field_value_sanitized[:5000] + '...'
            
            replacement_dict[field_name] = field_value_sanitized
            
            # Build rich context for this field
            field_context = f"• {field_label_ar} / {field_label_en}"
            
            # Add field type information
            type_mapping = {
                'text': 'نص / Text',
                'textarea': 'نص طويل / Long Text',
                'number': 'رقم / Number',
                'select': 'اختيار / Selection',
                'date': 'تاريخ / Date'
            }
            field_type_display = type_mapping.get(field_type, field_type)
            field_context += f" ({field_type_display})"
            
            # Add value
            field_context += f": {field_value_sanitized}"
            
            # Add available options for select fields
            if field_type == 'select' and field_options:
                options_str = ', '.join(field_options)
                field_context += f"\n  الخيارات المتاحة / Available options: [{options_str}]"
            
            input_context_parts.append(field_context)
        
        # Combine structured input context
        structured_input_context = "\n".join(input_context_parts) if input_context_parts else ""
        
        # Replace {field_name} with actual values in prompt template
        system_prompt = offering.ai_prompt_template
        for field_name, field_value in replacement_dict.items():
            # Use safe replacement - only replace exact {field_name} pattern
            placeholder = f'{{{field_name}}}'
            if placeholder in system_prompt:
                system_prompt = system_prompt.replace(placeholder, str(field_value))
        
        # إضافة تعليمات للذكاء الاصطناعي حول كيفية التعامل مع المدخلات
        input_handling_instructions = """

**تعليمات التعامل مع المدخلات (Input Handling Instructions):**
- اقرأ بعناية نوع كل حقل (Field Type) المذكور في بيانات المستخدم
- الحقول الرقمية (Number): استخدمها في الحسابات والتحليلات الكمية
- حقول الاختيار (Selection): افهم أن المستخدم اختار من بين عدة خيارات، وركز على الخيار المحدد
- النصوص الطويلة (Long Text): احلل المحتوى بتفصيل واستخرج النقاط الرئيسية
- التواريخ (Date): استخدمها في الجدولة الزمنية والتخطيط
- استخدم جميع المدخلات المتاحة بذكاء لتقديم استشارة دقيقة وشاملة

**تعليمات التنسيق (Formatting Instructions):**
- استخدم تنسيق Markdown للرد
- استخدم العناوين (#, ##, ###) لتنظيم المحتوى
- استخدم القوائم غير المرقمة (- ) للنقاط المهمة (3+ نقاط)
- استخدم القوائم المرقمة (1. ) للخطوات والمراحل
- استخدم الجداول (|) لعرض البيانات المنظمة
- استخدم **النص الغامق** للتأكيد على النقاط الهامة
- نسق الأرقام والإحصائيات بصيغة: "رقم - وصف" (مثال: "75% - معدل النجاح")"""
        
        system_prompt += input_handling_instructions
        
        # Build comprehensive user message with structured input context
        user_message_parts = [
            f"المشروع / Project: {form_data.get('project_name', 'غير محدد')}"
        ]
        
        if form_data.get('description'):
            user_message_parts.append(f"\nالوصف / Description:\n{form_data.get('description')}")
        
        if form_data.get('additional_context'):
            user_message_parts.append(f"\nمعلومات إضافية / Additional Context:\n{form_data.get('additional_context')}")
        
        # Add structured input context (field metadata)
        if structured_input_context:
            user_message_parts.append(f"\n**بيانات الحقول المدخلة / Input Field Data:**\n{structured_input_context}")
        
        user_message_parts.append("\nيرجى تقديم استشارة شاملة ومفصلة بناءً على جميع البيانات المتاحة أعلاه / Please provide comprehensive consultation based on all the data provided above.")
        
        user_message = "\n".join(user_message_parts)
    else:
        # Default prompt (fallback)
        system_prompt = f"""أنت مستشار خبير في {service.title_ar if lang == 'ar' else service.title_en}.
مهمتك تقديم استشارات احترافية وشاملة في مجال {offering.title_ar if lang == 'ar' else offering.title_en}.
قدم تحليلاً دقيقاً وتوصيات عملية بناءً على المعلومات المقدمة.

**تعليمات التنسيق:**
- استخدم تنسيق Markdown للرد
- استخدم العناوين (#, ##, ###) لتنظيم المحتوى
- استخدم القوائم للنقاط المهمة
- استخدم الجداول لعرض البيانات المنظمة
- نسق الأرقام والإحصائيات بوضوح"""
        
        user_message = f"""المشروع: {form_data.get('project_name', 'غير محدد')}

الوصف:
{form_data.get('description', '')}

{f"معلومات إضافية: {form_data.get('additional_context', '')}" if form_data.get('additional_context') else ''}

يرجى تقديم استشارة شاملة ومفصلة منسقة بـ Markdown."""
    
    try:
        # Use HuggingFace AI via AIManager (same as Strategic Planning)
        ai_manager = AIManager.for_use_case('custom_consultation')
        response_text = ai_manager.chat(system_prompt, user_message)
        
        # Log AI usage
        ai_log = AILog(
            user_id=int(user_id),
            module=f"{service_slug}_{offering_slug}",
            prompt=f"{system_prompt}\n\nUser: {user_message}",
            response=response_text,
            tokens_used=offering.ai_credits_cost or 1
        )
        db.session.add(ai_log)
        
        # Update user credits
        user.ai_credits_used += (offering.ai_credits_cost or 1)
        
        # Create project
        project = Project(
            user_id=int(user_id),
            title=f"{offering.title_ar if lang == 'ar' else offering.title_en} - {form_data.get('project_name', 'جديد')}",
            module=f"{service_slug}_{offering_slug}",
            content=json.dumps({
                'input': form_data,
                'output': response_text
            }, ensure_ascii=False),
            status='completed'
        )
        db.session.add(project)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'response': response_text,
            'project_id': project.id,
            'credits_used': offering.ai_credits_cost or 1,
            'credits_remaining': (plan.ai_credits_limit - user.ai_credits_used) if plan and plan.ai_credits_limit else None
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@services_bp.route('/project/<int:project_id>/view')
@login_required
def view_project(project_id):
    """View a consultation project"""
    from flask import abort
    
    db = get_db()
    lang = session.get('language', 'ar')
    
    # Get project
    project = db.session.get(Project, project_id)
    if not project:
        abort(404)
    
    # Check ownership or admin access
    user_id = get_jwt_identity()
    current_user = db.session.get(User, int(user_id))
    
    # Allow if user owns the project OR is system_admin
    if project.user_id != int(user_id) and current_user.role != 'system_admin':
        abort(403)
    
    # Parse module to get service and offering info
    service = None
    offering = None
    if project.module and '_' in project.module:
        parts = project.module.split('_', 1)
        if len(parts) == 2:
            service_slug, offering_slug = parts
            service = db.session.query(Service).filter_by(slug=service_slug).first()
            if service:
                offering = db.session.query(ServiceOffering).filter_by(
                    service_id=service.id,
                    slug=offering_slug
                ).first()
    
    # Parse project content
    project_data = {}
    try:
        project_data = json.loads(project.content) if project.content else {}
    except:
        project_data = {'input': {}, 'output': ''}
    
    return render_template(
        'services/project_view.html',
        project=project,
        service=service,
        offering=offering,
        project_data=project_data,
        lang=lang
    )

@services_bp.route('/project/<int:project_id>/export-pdf')
@login_required
def export_project_pdf(project_id):
    """تصدير المشروع كملف PDF مع التنسيق الكامل"""
    from flask import abort
    
    db = get_db()
    lang = session.get('language', 'ar')
    
    # Get project
    project = db.session.get(Project, project_id)
    if not project:
        abort(404)
    
    # Check ownership - get JWT identity and handle None
    user_id = get_jwt_identity()
    if user_id is None:
        abort(401)  # Unauthorized
    
    if project.user_id != int(user_id):
        abort(403)  # Forbidden
    
    # Parse project content
    project_data = {}
    try:
        project_data = json.loads(project.content) if project.content else {}
    except:
        project_data = {'input': {}, 'output': ''}
    
    # Get service and offering info
    service = None
    offering = None
    if project.module and '_' in project.module:
        parts = project.module.split('_', 1)
        if len(parts) == 2:
            service_slug, offering_slug = parts
            service = db.session.query(Service).filter_by(slug=service_slug).first()
            if service:
                offering = db.session.query(ServiceOffering).filter_by(
                    service_id=service.id,
                    slug=offering_slug
                ).first()
    
    try:
        # Generate HTML for PDF
        html_content = _generate_project_html(project, service, offering, project_data, lang)
        
        # Convert to PDF using WeasyPrint
        pdf_file = HTML(string=html_content).write_pdf()
        
        # Create response
        response = make_response(pdf_file)
        response.headers['Content-Type'] = 'application/pdf'
        filename = f"consultation_{project_id}.pdf"
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        current_app.logger.error(f"PDF export error: {str(e)}")
        abort(500)

@services_bp.route('/project/<int:project_id>/export-excel')
@login_required
def export_project_excel(project_id):
    """تصدير المشروع كملف Excel"""
    from flask import abort
    
    db = get_db()
    lang = session.get('language', 'ar')
    
    # Get project
    project = db.session.get(Project, project_id)
    if not project:
        abort(404)
    
    # Check ownership - get JWT identity and handle None
    user_id = get_jwt_identity()
    if user_id is None:
        abort(401)  # Unauthorized
    
    if project.user_id != int(user_id):
        abort(403)  # Forbidden
    
    # Parse project content
    project_data = {}
    try:
        project_data = json.loads(project.content) if project.content else {}
    except:
        project_data = {'input': {}, 'output': ''}
    
    # Get service and offering info
    service = None
    offering = None
    if project.module and '_' in project.module:
        parts = project.module.split('_', 1)
        if len(parts) == 2:
            service_slug, offering_slug = parts
            service = db.session.query(Service).filter_by(slug=service_slug).first()
            if service:
                offering = db.session.query(ServiceOffering).filter_by(
                    service_id=service.id,
                    slug=offering_slug
                ).first()
    
    try:
        from utils.markdown_formatter import extract_sections_from_markdown, clean_markdown_for_excel
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "الاستشارة" if lang == 'ar' else "Consultation"
        ws.sheet_view.rightToLeft = (lang == 'ar')
        
        # Get service color
        service_color = service.color.replace('#', '') if service and service.color else '1e3a8a'
        
        # Define professional styles
        title_fill = PatternFill(start_color=service_color, end_color=service_color, fill_type="solid")
        title_font = Font(name='Arial', size=18, bold=True, color="FFFFFF")
        
        header_fill = PatternFill(start_color=service_color, end_color=service_color, fill_type="solid")
        header_font = Font(name='Arial', size=14, bold=True, color="FFFFFF")
        
        section_fill = PatternFill(start_color="e2e8f0", end_color="e2e8f0", fill_type="solid")
        section_font = Font(name='Arial', size=12, bold=True, color="2c5282")
        
        key_fill = PatternFill(start_color="f7fafc", end_color="f7fafc", fill_type="solid")
        key_font = Font(name='Arial', size=10, bold=True, color="2c5282")
        value_font = Font(name='Arial', size=10)
        
        normal_font = Font(name='Arial', size=10)
        rtl_alignment = Alignment(horizontal='right', vertical='top', wrap_text=True, readingOrder=2)
        ltr_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        row = 1
        
        # Title
        ws[f'A{row}'] = project.title
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].fill = title_fill
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'A{row}:E{row}')
        ws.row_dimensions[row].height = 35
        row += 1
        
        # Service info
        if service and offering:
            ws[f'A{row}'] = f"{service.title_ar if lang == 'ar' else service.title_en} - {offering.title_ar if lang == 'ar' else offering.title_en}"
            ws[f'A{row}'].font = Font(name='Arial', size=11, color="4a5568")
            ws[f'A{row}'].alignment = Alignment(horizontal='center')
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
        
        # Date
        ws[f'A{row}'] = f"{'التاريخ' if lang == 'ar' else 'Date'}: {project.created_at.strftime('%Y-%m-%d %H:%M')}"
        ws[f'A{row}'].font = Font(name='Arial', size=10, color="718096")
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        ws.merge_cells(f'A{row}:E{row}')
        row += 2
        
        # Input section header
        ws[f'A{row}'] = "البيانات المدخلة" if lang == 'ar' else "Input Data"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'A{row}:E{row}')
        ws.row_dimensions[row].height = 30
        row += 1
        
        # Input data
        input_data = project_data.get('input', {})
        for key, value in input_data.items():
            ws[f'A{row}'] = str(key).replace('_', ' ').title()
            ws[f'A{row}'].font = key_font
            ws[f'A{row}'].fill = key_fill
            ws[f'A{row}'].alignment = rtl_alignment if lang == 'ar' else ltr_alignment
            
            ws[f'B{row}'] = str(value)
            ws[f'B{row}'].font = value_font
            ws[f'B{row}'].alignment = rtl_alignment if lang == 'ar' else ltr_alignment
            ws.merge_cells(f'B{row}:E{row}')
            ws.row_dimensions[row].height = 20
            row += 1
        
        row += 1
        
        # Output section header
        ws[f'A{row}'] = "نتيجة الاستشارة الذكية" if lang == 'ar' else "AI Consultation Result"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'A{row}:E{row}')
        ws.row_dimensions[row].height = 30
        row += 1
        
        # Output content with sections
        output_text = project_data.get('output', '')
        sections = extract_sections_from_markdown(output_text)
        
        if sections:
            for section in sections:
                # Section title
                ws[f'A{row}'] = section['title']
                ws[f'A{row}'].font = section_font
                ws[f'A{row}'].fill = section_fill
                ws[f'A{row}'].alignment = rtl_alignment if lang == 'ar' else ltr_alignment
                ws.merge_cells(f'A{row}:E{row}')
                ws.row_dimensions[row].height = 25
                row += 1
                
                # Section content
                content = clean_markdown_for_excel(section['content'])
                ws[f'A{row}'] = content
                ws[f'A{row}'].font = normal_font
                ws[f'A{row}'].alignment = rtl_alignment if lang == 'ar' else ltr_alignment
                ws.merge_cells(f'A{row}:E{row}')
                
                lines = content.count('\n') + 1
                ws.row_dimensions[row].height = max(20, min(lines * 15, 200))
                row += 2
        else:
            content = clean_markdown_for_excel(output_text)
            ws[f'A{row}'] = content
            ws[f'A{row}'].font = normal_font
            ws[f'A{row}'].alignment = rtl_alignment if lang == 'ar' else ltr_alignment
            ws.merge_cells(f'A{row}:E{row}')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        filename = f"consultation_{project_id}.xlsx"
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        current_app.logger.error(f"Excel export error: {str(e)}")
        abort(500)

def _generate_project_html(project, service, offering, project_data, lang):
    """Generate HTML for PDF export with professionally formatted consultation content"""
    from utils.markdown_formatter import format_consultation_output
    
    input_data = project_data.get('input', {})
    output_text = project_data.get('output', '')
    
    # Format Markdown to beautiful HTML with cards, grids, tables, stat boxes
    formatted_output = format_consultation_output(output_text, lang)
    
    # Render the professional export template
    html = render_template(
        'exports/project_export.html',
        project=project,
        service=service,
        offering=offering,
        input_data=input_data,
        formatted_output=formatted_output,
        lang=lang
    )
    
    return html
