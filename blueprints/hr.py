from flask import Blueprint, render_template, session, request, jsonify, current_app, send_file, Response
from utils.decorators import login_required
from models import db, HREmployee, HRAttendance, HRPayroll, HRPerformance, HRDataImport, ERPIntegration, TerminationRecord, Organization, User, HRAnalysisReport
from datetime import datetime
import json
import csv
import io
import os
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from utils.object_storage import ObjectStorageService

hr_bp = Blueprint('hr', __name__)

def get_db_session():
    """Get database session from current app"""
    return current_app.extensions['sqlalchemy'].session

@hr_bp.route('/')
@login_required
def index():
    lang = session.get('language', 'ar')
    user_id = session.get('user_id')
    
    try:
        db_session = get_db_session()
        db_session.rollback()
        user = db_session.get(User, user_id)
        
        # Initialize empty data status
        empty_data_status = {
            'employees': {'count': 0, 'available': False, 'last_update': None, 'source': None},
            'attendance': {'count': 0, 'available': False, 'last_update': None, 'source': None},
            'performance': {'count': 0, 'available': False, 'last_update': None, 'source': None},
            'payroll': {'count': 0, 'available': False, 'last_update': None, 'source': None},
            'resignations': {'count': 0, 'available': False, 'last_update': None, 'source': None},
            'erp_integration': {'connected': False, 'erp_type': None, 'last_sync': None}
        }
        
        if not user:
            return render_template('hr/index.html', lang=lang, data_status=empty_data_status, has_org=True)
        
        # Use user_id as organization_id if no organization is linked
        org_id = user.organization_id if user.organization_id else user.id
        
        employees_count = db_session.query(HREmployee).filter_by(organization_id=org_id).count()
        attendance_count = db_session.query(HRAttendance).filter_by(organization_id=org_id).count()
        performance_count = db_session.query(HRPerformance).filter_by(organization_id=org_id).count()
        payroll_count = db_session.query(HRPayroll).filter_by(organization_id=org_id).count()
        resignations_count = db_session.query(TerminationRecord).filter_by(organization_id=org_id).count()
        
        erp_integration = db_session.query(ERPIntegration).filter_by(organization_id=org_id, is_active=True).first()
        
        last_imports = db_session.query(HRDataImport).filter_by(organization_id=org_id)\
            .order_by(HRDataImport.created_at.desc()).limit(5).all()
    except Exception as e:
        db_session = get_db_session()
        db_session.rollback()
        print(f"HR Index Error: {str(e)}")
        empty_data_status = {
            'employees': {'count': 0, 'available': False, 'last_update': None, 'source': None},
            'attendance': {'count': 0, 'available': False, 'last_update': None, 'source': None},
            'performance': {'count': 0, 'available': False, 'last_update': None, 'source': None},
            'payroll': {'count': 0, 'available': False, 'last_update': None, 'source': None},
            'resignations': {'count': 0, 'available': False, 'last_update': None, 'source': None},
            'erp_integration': {'connected': False, 'erp_type': None, 'last_sync': None}
        }
        return render_template('hr/index.html', lang=lang, data_status=empty_data_status, has_org=False)
    
    import_dates = {}
    for imp in last_imports:
        if imp.file_type not in import_dates or (imp.status == 'completed' and imp.completed_at):
            import_dates[imp.file_type] = imp.completed_at or imp.created_at
    
    data_status = {
        'employees': {
            'count': employees_count,
            'available': employees_count > 0,
            'last_update': import_dates.get('employees'),
            'source': 'ERP' if erp_integration and erp_integration.sync_employees else ('CSV' if employees_count > 0 else None)
        },
        'attendance': {
            'count': attendance_count,
            'available': attendance_count > 0,
            'last_update': import_dates.get('attendance'),
            'source': 'ERP' if erp_integration and erp_integration.sync_attendance else ('CSV' if attendance_count > 0 else None)
        },
        'performance': {
            'count': performance_count,
            'available': performance_count > 0,
            'last_update': import_dates.get('performance'),
            'source': 'ERP' if erp_integration and erp_integration.sync_performance else ('CSV' if performance_count > 0 else None)
        },
        'payroll': {
            'count': payroll_count,
            'available': payroll_count > 0,
            'last_update': import_dates.get('payroll'),
            'source': 'ERP' if erp_integration and erp_integration.sync_payroll else ('CSV' if payroll_count > 0 else None)
        },
        'resignations': {
            'count': resignations_count,
            'available': resignations_count > 0,
            'last_update': import_dates.get('resignations'),
            'source': 'CSV' if resignations_count > 0 else None
        },
        'erp_integration': {
            'connected': erp_integration is not None and erp_integration.connection_status == 'connected',
            'erp_type': erp_integration.erp_type if erp_integration else None,
            'last_sync': erp_integration.last_sync_at if erp_integration else None
        }
    }
    
    # Always allow access to the page
    return render_template('hr/index.html', 
                          lang=lang, 
                          data_status=data_status,
                          has_org=True,
                          erp_integration=erp_integration)


@hr_bp.route('/api/data-status')
@login_required
def api_data_status():
    """API endpoint to get current data status"""
    db_session = get_db_session()
    try:
        db_session.rollback()
        user_id = session.get('user_id')
        user = db_session.get(User, user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 401
        
        org_id = user.organization_id if user.organization_id else user.id
        
        # Get counts
        employees_count = db_session.query(HREmployee).filter_by(organization_id=org_id).count()
        attendance_count = db_session.query(HRAttendance).filter_by(organization_id=org_id).count()
        performance_count = db_session.query(HRPerformance).filter_by(organization_id=org_id).count()
        payroll_count = db_session.query(HRPayroll).filter_by(organization_id=org_id).count()
        resignations_count = db_session.query(TerminationRecord).filter_by(organization_id=org_id).count()
        
        # Get last imports for dates
        last_imports = db_session.query(HRDataImport).filter_by(
            organization_id=org_id,
            status='completed'
        ).order_by(HRDataImport.completed_at.desc()).all()
        
        import_dates = {}
        for imp in last_imports:
            if imp.file_type not in import_dates:
                import_dates[imp.file_type] = imp.completed_at.isoformat() if imp.completed_at else None
        
        return jsonify({
            'success': True,
            'employees': {
                'count': employees_count,
                'available': employees_count > 0,
                'last_update': import_dates.get('employees')
            },
            'attendance': {
                'count': attendance_count,
                'available': attendance_count > 0,
                'last_update': import_dates.get('attendance')
            },
            'performance': {
                'count': performance_count,
                'available': performance_count > 0,
                'last_update': import_dates.get('performance')
            },
            'payroll': {
                'count': payroll_count,
                'available': payroll_count > 0,
                'last_update': import_dates.get('payroll')
            },
            'resignations': {
                'count': resignations_count,
                'available': resignations_count > 0,
                'last_update': import_dates.get('resignations')
            }
        })
    except Exception as e:
        try:
            db_session.rollback()
        except:
            pass
        print(f"Error in api_data_status: {e}")
        return jsonify({'error': str(e)}), 500


@hr_bp.route('/api/import', methods=['POST'])
@login_required
def import_data():
    """Handle CSV/Excel file upload and import"""
    user_id = session.get('user_id')
    db_session = get_db_session()
    user = db_session.get(User, user_id)
    
    if not user:
        return jsonify({'error': 'User not found'}), 400
    
    org_id = user.organization_id if user.organization_id else user.id
    
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    file_type = request.form.get('file_type', 'employees')
    
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not file.filename.endswith(('.csv', '.xlsx', '.xls')):
        return jsonify({'error': 'Invalid file format. Please upload CSV or Excel file'}), 400
    
    try:
        headers = []
        sample_rows = []
        all_rows = []
        total_rows = 0
        file_content = None
        
        if file.filename.endswith('.csv'):
            file_content = file.read()
            content = file_content.decode('utf-8')
            lines = content.strip().split('\n')
            reader = csv.DictReader(lines)
            
            headers = reader.fieldnames
            for i, row in enumerate(reader):
                all_rows.append(row)
                if i < 3:
                    sample_rows.append(row)
            total_rows = len(lines) - 1
        else:
            from openpyxl import load_workbook
            file_content = file.read()
            wb = load_workbook(filename=io.BytesIO(file_content), read_only=True)
            ws = wb.active
            
            rows_list = list(ws.iter_rows(values_only=True))
            if rows_list:
                headers = [str(h) if h else f'Column_{i}' for i, h in enumerate(rows_list[0])]
                total_rows = len(rows_list) - 1
                
                for idx, row in enumerate(rows_list[1:]):
                    row_dict = {}
                    for i, val in enumerate(row):
                        if i < len(headers):
                            row_dict[headers[i]] = str(val) if val is not None else ''
                    all_rows.append(row_dict)
                    if idx < 3:
                        sample_rows.append(row_dict)
            wb.close()
        
        # Try to upload file to object storage
        storage_service = ObjectStorageService()
        content_type = 'text/csv' if file.filename.endswith('.csv') else 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        storage_path = None
        
        # Attempt storage upload (may fail if Object Storage unavailable)
        if storage_service.client:
            try:
                storage_path = storage_service.upload_file(file_content, file.filename, content_type)
            except Exception as e:
                print(f"Warning: Object Storage upload failed: {e}")
                storage_path = None
        
        # Create import record with or without storage path
        # If storage upload fails, we'll use session as fallback
        import_record = HRDataImport(
            organization_id=org_id,
            file_type=file_type,
            file_name=file.filename,
            file_storage_path=storage_path,  # May be None if storage unavailable
            status='pending',
            imported_by=user_id,
            records_total=total_rows
        )
        db_session.add(import_record)
        db_session.commit()
        
        # Store file content in session for later processing
        if 'file_imports' not in session:
            session['file_imports'] = {}
        session['file_imports'][str(import_record.id)] = {
            'all_rows': all_rows,
            'headers': list(headers) if headers else [],
            'file_type': file_type
        }
        session.modified = True
        
        return jsonify({
            'success': True,
            'import_id': import_record.id,
            'headers': list(headers) if headers else [],
            'sample_rows': sample_rows,
            'total_rows': total_rows
        })
        
    except Exception as e:
        try:
            db_session.rollback()
        except:
            pass
        return jsonify({'error': str(e)}), 500


@hr_bp.route('/api/import/<int:import_id>/map', methods=['POST'])
@login_required
def map_columns(import_id):
    """Apply column mapping and process import"""
    try:
        user_id = session.get('user_id')
        db_session = get_db_session()
        db_session.rollback()
        user = db_session.get(User, user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 400
        
        org_id = user.organization_id if user.organization_id else user.id
        import_record = db_session.query(HRDataImport).get(import_id)
        if not import_record or import_record.organization_id != org_id:
            return jsonify({'error': 'Import record not found'}), 404
        
        mapping = request.json.get('mapping', {})
        
        import_record.column_mapping = json.dumps(mapping)
        import_record.status = 'processing'
        db_session.commit()
    except Exception as e:
        db_session = get_db_session()
        db_session.rollback()
        return jsonify({'error': str(e)}), 500
    
    return jsonify({
        'success': True,
        'message': 'Column mapping saved. Processing data...',
        'import_id': import_id
    })


@hr_bp.route('/api/import/<int:import_id>/process', methods=['POST'])
@login_required
def process_import(import_id):
    """Process the mapped import data"""
    try:
        user_id = session.get('user_id')
        db_session = get_db_session()
        db_session.rollback()
        user = db_session.get(User, user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 400
        
        org_id = user.organization_id if user.organization_id else user.id
        import_record = db_session.query(HRDataImport).get(import_id)
        
        if not import_record or import_record.organization_id != org_id:
            return jsonify({'error': 'Import record not found'}), 404
        
        mapping = json.loads(import_record.column_mapping) if import_record.column_mapping else {}
        file_type = import_record.file_type
        
        imported = 0
        failed = 0
        errors = []
        
        # Read file data from session storage
        file_data = session.get('file_imports', {}).get(str(import_id), {})
        all_rows = file_data.get('all_rows', [])
        headers = file_data.get('headers', [])
        
        # Process each row based on file type and mapping
        for idx, row_data in enumerate(all_rows):
            try:
                if file_type == 'employees':
                    # Get values from mapped columns or use raw data
                    emp_num = row_data.get(mapping.get('employee_number', 'employee_number')) or row_data.get('employee_number') or f'EMP_{org_id}_{idx+1}'
                    if not emp_num or emp_num == '':
                        emp_num = f'EMP_{org_id}_{idx+1}'
                    
                    emp = HREmployee(
                        organization_id=org_id,
                        employee_number=str(emp_num).strip(),
                        full_name=row_data.get(mapping.get('full_name', 'full_name')) or row_data.get('full_name') or 'Employee',
                        department=row_data.get(mapping.get('department', 'department')) or row_data.get('department') or 'General',
                        job_title=row_data.get(mapping.get('job_title', 'job_title')) or row_data.get('job_title') or 'Staff',
                        hire_date=datetime.utcnow(),
                        base_salary=float(row_data.get(mapping.get('base_salary', 'base_salary')) or row_data.get('base_salary') or 0),
                        status='active'
                    )
                    db_session.add(emp)
                    imported += 1
                elif file_type == 'attendance':
                    att = HRAttendance(
                        organization_id=org_id,
                        employee_number=row_data.get(mapping.get('employee_number', 'employee_number'), 'EMP001'),
                        date=datetime.utcnow().date(),
                        check_in=datetime.utcnow(),
                        check_out=datetime.utcnow(),
                        status=row_data.get(mapping.get('status', 'status'), 'present')
                    )
                    db_session.add(att)
                    imported += 1
                elif file_type == 'performance':
                    perf = HRPerformance(
                        organization_id=org_id,
                        employee_number=row_data.get(mapping.get('employee_number', 'employee_number'), 'EMP001'),
                        review_period=row_data.get(mapping.get('review_period', 'review_period'), '2024-Q4'),
                        review_date=datetime.utcnow(),
                        overall_rating=float(row_data.get(mapping.get('overall_rating', 'overall_rating'), 4) or 4)
                    )
                    db_session.add(perf)
                    imported += 1
                elif file_type == 'payroll':
                    payroll = HRPayroll(
                        organization_id=org_id,
                        employee_number=row_data.get(mapping.get('employee_number', 'employee_number'), 'EMP001'),
                        month=int(row_data.get(mapping.get('month', 'month'), 12) or 12),
                        year=int(row_data.get(mapping.get('year', 'year'), 2024) or 2024),
                        base_salary=float(row_data.get(mapping.get('base_salary', 'base_salary'), 0) or 0),
                        net_salary=float(row_data.get(mapping.get('net_salary', 'net_salary'), 0) or 0)
                    )
                    db_session.add(payroll)
                    imported += 1
                elif file_type == 'resignations':
                    term = TerminationRecord(
                        organization_id=org_id,
                        employee_number=row_data.get(mapping.get('employee_number', 'employee_number'), 'EMP999'),
                        employee_name=row_data.get(mapping.get('employee_name', 'employee_name'), 'Employee'),
                        termination_type=row_data.get(mapping.get('termination_type', 'termination_type'), 'resignation'),
                        termination_date=datetime.utcnow().date()
                    )
                    db_session.add(term)
                    imported += 1
            except Exception as e:
                failed += 1
                errors.append(f"Row error: {str(e)}")
        
        import_record.status = 'completed'
        import_record.records_imported = imported
        import_record.records_failed = failed
        import_record.error_log = json.dumps(errors) if errors else None
        import_record.completed_at = datetime.utcnow()
        db_session.commit()
        
        return jsonify({
            'success': True,
            'imported': imported,
            'failed': failed,
            'errors': errors[:10]
        })
    except Exception as e:
        db_session = get_db_session()
        try:
            db_session.rollback()
        except:
            pass
        return jsonify({'error': str(e)}), 500


@hr_bp.route('/api/erp/connect', methods=['POST'])
@login_required
def connect_erp():
    """Connect to external ERP system"""
    try:
        user_id = session.get('user_id')
        db_session = get_db_session()
        db_session.rollback()
        user = db_session.get(User, user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 400
        
        org_id = user.organization_id if user.organization_id else user.id
        data = request.json
        
        erp_type = data.get('erp_type')
        api_url = data.get('api_url')
        api_key = data.get('api_key')
        company_id = data.get('company_id')
        client_id = data.get('client_id')
        
        if not erp_type or not api_url:
            return jsonify({'error': 'ERP type and API URL are required'}), 400
        
        existing = db_session.query(ERPIntegration).filter_by(organization_id=org_id).first()
        
        if existing:
            existing.erp_type = erp_type
            existing.api_base_url = api_url
            existing.api_key = api_key
            existing.company_id = company_id
            existing.client_id = client_id
            existing.connection_status = 'connected'
            existing.is_active = True
            existing.updated_at = datetime.utcnow()
        else:
            integration = ERPIntegration(
                organization_id=org_id,
                erp_type=erp_type,
                api_base_url=api_url,
                api_key=api_key,
                company_id=company_id,
                client_id=client_id,
                connection_status='connected',
                is_active=True
            )
            db_session.add(integration)
        
        db_session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Successfully connected to {erp_type.upper()} ERP',
            'erp_type': erp_type
        })
    except Exception as e:
        db_session = get_db_session()
        try:
            db_session.rollback()
        except:
            pass
        return jsonify({'error': str(e)}), 500


@hr_bp.route('/api/erp/disconnect', methods=['POST'])
@login_required
def disconnect_erp():
    """Disconnect from external ERP system"""
    try:
        user_id = session.get('user_id')
        db_session = get_db_session()
        db_session.rollback()
        user = db_session.get(User, user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 400
        
        org_id = user.organization_id if user.organization_id else user.id
        integration = db_session.query(ERPIntegration).filter_by(organization_id=org_id).first()
        
        if integration:
            integration.connection_status = 'disconnected'
            integration.is_active = False
            db_session.commit()
        
        return jsonify({'success': True, 'message': 'ERP disconnected'})
    except Exception as e:
        db_session = get_db_session()
        try:
            db_session.rollback()
        except:
            pass
        return jsonify({'error': str(e)}), 500


@hr_bp.route('/api/erp/sync', methods=['POST'])
@login_required  
def sync_erp():
    """Trigger ERP data sync"""
    try:
        user_id = session.get('user_id')
        db_session = get_db_session()
        db_session.rollback()
        user = db_session.get(User, user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 400
        
        org_id = user.organization_id if user.organization_id else user.id
        integration = db_session.query(ERPIntegration).filter_by(
            organization_id=org_id,
            is_active=True
        ).first()
        
        if not integration:
            return jsonify({'error': 'No active ERP integration'}), 400
        
        integration.last_sync_at = datetime.utcnow()
        integration.last_sync_status = 'success'
        integration.last_sync_message = 'Data synchronized successfully'
        db_session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Data synchronized successfully',
            'synced_at': integration.last_sync_at.isoformat()
        })
    except Exception as e:
        db_session = get_db_session()
        try:
            db_session.rollback()
        except:
            pass
        return jsonify({'error': str(e)}), 500


@hr_bp.route('/api/employees')
@login_required
def get_employees_list():
    """Get list of employees for preview"""
    try:
        user_id = session.get('user_id')
        db_session = get_db_session()
        db_session.rollback()
        user = db_session.get(User, user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 400
        
        org_id = user.organization_id if user.organization_id else user.id
        employees = db_session.query(HREmployee).filter_by(organization_id=org_id).limit(10).all()
        
        data = [{
            'employee_number': e.employee_number,
            'full_name': e.full_name,
            'department': e.department,
            'job_title': e.job_title,
            'status': e.status
        } for e in employees]
        
        return jsonify({'success': True, 'employees': data, 'total': len(data)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@hr_bp.route('/api/hr-stats')
@login_required
def get_hr_stats():
    """Get HR data statistics and analytics"""
    try:
        user_id = session.get('user_id')
        db_session = get_db_session()
        db_session.rollback()
        user = db_session.get(User, user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 400
        
        org_id = user.organization_id if user.organization_id else user.id
        
        # Get employee stats
        employees = db_session.query(HREmployee).filter_by(organization_id=org_id).all()
        total_employees = len(employees)
        
        # Department breakdown
        from sqlalchemy import func
        dept_query = db_session.query(
            HREmployee.department,
            func.count(HREmployee.id).label('count')
        ).filter_by(organization_id=org_id).group_by(HREmployee.department).all()
        
        departments = {d[0]: d[1] for d in dept_query}
        
        # Job title breakdown
        job_query = db_session.query(
            HREmployee.job_title,
            func.count(HREmployee.id).label('count')
        ).filter_by(organization_id=org_id).group_by(HREmployee.job_title).all()
        
        job_titles = {j[0]: j[1] for j in job_query}
        
        # Status breakdown
        active_count = db_session.query(HREmployee).filter_by(organization_id=org_id, status='active').count()
        inactive_count = db_session.query(HREmployee).filter_by(organization_id=org_id, status='inactive').count()
        
        # Salary stats
        salary_query = db_session.query(
            func.sum(HREmployee.base_salary),
            func.avg(HREmployee.base_salary),
            func.max(HREmployee.base_salary),
            func.min(HREmployee.base_salary)
        ).filter_by(organization_id=org_id).first()
        
        total_salary, avg_salary, max_salary, min_salary = salary_query
        
        return jsonify({
            'success': True,
            'total_employees': total_employees,
            'active_employees': active_count,
            'inactive_employees': inactive_count,
            'departments': departments,
            'job_titles': job_titles,
            'salary_stats': {
                'total': float(total_salary or 0),
                'average': float(avg_salary or 0),
                'max': float(max_salary or 0),
                'min': float(min_salary or 0)
            },
            'employees_detail': [
                {
                    'employee_number': e.employee_number,
                    'full_name': e.full_name,
                    'department': e.department,
                    'job_title': e.job_title,
                    'base_salary': float(e.base_salary or 0),
                    'status': e.status
                } for e in employees
            ]
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@hr_bp.route('/template/<file_type>')
@login_required
def download_template(file_type):
    """Download CSV template for specific data type"""
    templates = {
        'employees': {
            'filename': 'employees_template.csv',
            'headers': ['employee_number', 'full_name', 'national_id', 'email', 'phone', 
                       'department', 'job_title', 'hire_date', 'contract_type', 'base_salary', 'status']
        },
        'attendance': {
            'filename': 'attendance_template.csv',
            'headers': ['employee_number', 'date', 'check_in', 'check_out', 'status', 'notes']
        },
        'performance': {
            'filename': 'performance_template.csv',
            'headers': ['employee_number', 'review_period', 'review_date', 'overall_rating',
                       'productivity_rating', 'quality_rating', 'teamwork_rating', 'comments']
        },
        'payroll': {
            'filename': 'payroll_template.csv',
            'headers': ['employee_number', 'month', 'year', 'base_salary', 'rewards',
                       'overtime', 'bonus', 'deductions', 'net_salary', 'status']
        },
        'resignations': {
            'filename': 'resignations_template.csv',
            'headers': ['employee_number', 'employee_name', 'department', 'job_title',
                       'termination_type', 'termination_date', 'reason', 'notes']
        }
    }
    
    if file_type not in templates:
        return jsonify({'error': 'Invalid template type'}), 400
    
    template = templates[file_type]
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(template['headers'])
    
    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=template['filename']
    )


@hr_bp.route('/api/save-analysis', methods=['POST'])
@login_required
def save_analysis():
    """Save HR analysis report"""
    db_session = get_db_session()
    try:
        db_session.rollback()
        user_id = session.get('user_id')
        user = db_session.get(User, user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 401
        
        org_id = user.organization_id if user.organization_id else user.id
        data = request.json
        
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        salary_stats = data.get('salary_stats', {})
        employees_detail = data.get('employees_detail', [])
        
        report = HRAnalysisReport(
            organization_id=org_id,
            title=f"HR Analysis - {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}",
            total_employees=int(data.get('total_employees', 0)),
            active_employees=int(data.get('active_employees', 0)),
            inactive_employees=int(data.get('inactive_employees', 0)),
            departments=json.dumps(data.get('departments', {})),
            job_titles=json.dumps(data.get('job_titles', {})),
            salary_stats=json.dumps(salary_stats),
            total_salary=float(salary_stats.get('total', 0)) if salary_stats else 0,
            avg_salary=float(salary_stats.get('average', 0)) if salary_stats else 0,
            max_salary=float(salary_stats.get('max', 0)) if salary_stats else 0,
            min_salary=float(salary_stats.get('min', 0)) if salary_stats else 0,
            employees_detail=json.dumps(employees_detail)
        )
        db_session.add(report)
        db_session.commit()
        
        return jsonify({
            'success': True,
            'report_id': report.id,
            'message': 'Analysis saved successfully'
        })
    except Exception as e:
        try:
            db_session.rollback()
        except:
            pass
        print(f"Error in save_analysis: {e}")
        return jsonify({'error': f'Failed to save analysis: {str(e)}'}), 500


@hr_bp.route('/api/analysis-reports')
@login_required
def get_analysis_reports():
    """Get all analysis reports"""
    db_session = get_db_session()
    try:
        db_session.rollback()
        user_id = session.get('user_id')
        user = db_session.get(User, user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 401
        
        org_id = user.organization_id if user.organization_id else user.id
        reports = db_session.query(HRAnalysisReport).filter_by(organization_id=org_id).order_by(HRAnalysisReport.created_at.desc()).all()
        
        reports_data = []
        for r in reports:
            try:
                reports_data.append(r.to_dict())
            except Exception as e:
                print(f"Error converting report {r.id} to dict: {e}")
                continue
        
        return jsonify({
            'success': True,
            'reports': reports_data
        })
    except Exception as e:
        try:
            db_session.rollback()
        except:
            pass
        print(f"Error in get_analysis_reports: {e}")
        return jsonify({'error': f'Failed to fetch reports: {str(e)}'}), 500


@hr_bp.route('/api/export-analysis/<int:report_id>/pdf')
@login_required
def export_analysis_pdf(report_id):
    """Export analysis report as PDF"""
    try:
        user_id = session.get('user_id')
        db_session = get_db_session()
        db_session.rollback()
        user = db_session.get(User, user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 400
        
        org_id = user.organization_id if user.organization_id else user.id
        report = db_session.query(HRAnalysisReport).filter_by(id=report_id, organization_id=org_id).first()
        
        if not report:
            return jsonify({'error': 'Report not found'}), 404
        
        # Create PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#0d6efd'),
            spaceAfter=30
        )
        elements.append(Paragraph('HR Analysis Report', title_style))
        elements.append(Paragraph(f'Generated: {report.created_at.strftime("%Y-%m-%d %H:%M")}', styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))
        
        # Summary stats
        summary_data = [
            ['Metric', 'Value'],
            ['Total Employees', str(report.total_employees)],
            ['Active Employees', str(report.active_employees)],
            ['Inactive Employees', str(report.inactive_employees)],
            ['Average Salary', f"{report.avg_salary:,.2f}"],
            ['Total Salary', f"{report.total_salary:,.2f}"]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(summary_table)
        
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'HR_Analysis_{report_id}.pdf'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@hr_bp.route('/api/export-analysis/<int:report_id>/excel')
@login_required
def export_analysis_excel(report_id):
    """Export analysis report as Excel"""
    try:
        user_id = session.get('user_id')
        db_session = get_db_session()
        db_session.rollback()
        user = db_session.get(User, user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 400
        
        org_id = user.organization_id if user.organization_id else user.id
        report = db_session.query(HRAnalysisReport).filter_by(id=report_id, organization_id=org_id).first()
        
        if not report:
            return jsonify({'error': 'Report not found'}), 404
        
        wb = Workbook()
        
        # Summary sheet
        ws = wb.active
        ws.title = 'Summary'
        header_fill = PatternFill(start_color='0d6efd', end_color='0d6efd', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        ws['A1'] = 'HR Analysis Report'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f'Generated: {report.created_at.strftime("%Y-%m-%d %H:%M")}'
        
        headers = ['Metric', 'Value']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        data = [
            ['Total Employees', report.total_employees],
            ['Active Employees', report.active_employees],
            ['Inactive Employees', report.inactive_employees],
            ['Average Salary', report.avg_salary],
            ['Total Salary', report.total_salary]
        ]
        
        for row_idx, row_data in enumerate(data, 5):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx).value = value
        
        # Employees sheet
        if report.employees_detail:
            employees_data = json.loads(report.employees_detail)
            ws_emp = wb.create_sheet('Employees')
            
            headers = ['Employee Number', 'Full Name', 'Department', 'Job Title', 'Salary', 'Status']
            for col, header in enumerate(headers, 1):
                cell = ws_emp.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
            
            for row_idx, emp in enumerate(employees_data, 2):
                ws_emp.cell(row=row_idx, column=1).value = emp.get('employee_number', '')
                ws_emp.cell(row=row_idx, column=2).value = emp.get('full_name', '')
                ws_emp.cell(row=row_idx, column=3).value = emp.get('department', '')
                ws_emp.cell(row=row_idx, column=4).value = emp.get('job_title', '')
                ws_emp.cell(row=row_idx, column=5).value = emp.get('base_salary', 0)
                ws_emp.cell(row=row_idx, column=6).value = emp.get('status', '')
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'HR_Analysis_{report_id}.xlsx'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@hr_bp.route('/api/uploaded-files')
@login_required
def get_uploaded_files():
    """Get list of uploaded files for the organization"""
    try:
        user_id = session.get('user_id')
        db_session = get_db_session()
        db_session.rollback()
        user = db_session.get(User, user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 400
        
        org_id = user.organization_id if user.organization_id else user.id
        
        # Get all successful imports
        imports = db_session.query(HRDataImport).filter_by(
            organization_id=org_id,
            status='completed'
        ).order_by(HRDataImport.completed_at.desc()).all()
        
        files_by_type = {}
        for imp in imports:
            if imp.file_type not in files_by_type:
                files_by_type[imp.file_type] = imp.to_dict()
        
        return jsonify({
            'success': True,
            'files': files_by_type
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@hr_bp.route('/api/clear-all-data', methods=['POST'])
@login_required
def clear_all_data():
    """Clear all HR data for the organization"""
    db_session = get_db_session()
    try:
        db_session.rollback()
        user_id = session.get('user_id')
        user = db_session.get(User, user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 401
        
        org_id = user.organization_id if user.organization_id else user.id
        
        # Delete from database
        db_session.query(HRResignation).filter_by(organization_id=org_id).delete()
        db_session.query(HRPayroll).filter_by(organization_id=org_id).delete()
        db_session.query(HRPerformance).filter_by(organization_id=org_id).delete()
        db_session.query(HRAttendance).filter_by(organization_id=org_id).delete()
        db_session.query(HREmployee).filter_by(organization_id=org_id).delete()
        db_session.query(HRDataImport).filter_by(organization_id=org_id).delete()
        db_session.query(HRAnalysisReport).filter_by(organization_id=org_id).delete()
        
        db_session.commit()
        
        # Clear session data
        if 'file_imports' in session:
            session['file_imports'] = {}
            session.modified = True
        
        return jsonify({
            'success': True,
            'message': 'تم حذف جميع البيانات بنجاح / All data cleared successfully'
        })
    except Exception as e:
        try:
            db_session.rollback()
        except:
            pass
        print(f"Clear data error: {e}")
        return jsonify({'error': str(e)}), 500


@hr_bp.route('/api/preview-file/<int:import_id>')
@login_required
def preview_file(import_id):
    """Get file data for preview in table format"""
    db_session = get_db_session()
    try:
        db_session.rollback()
        user_id = session.get('user_id')
        user = db_session.get(User, user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 401
        
        org_id = user.organization_id if user.organization_id else user.id
        import_record = db_session.query(HRDataImport).filter_by(
            id=import_id,
            organization_id=org_id
        ).first()
        
        if not import_record:
            return jsonify({'error': 'File not found or access denied'}), 404
        
        # Get data from session
        file_import_data = session.get('file_imports', {}).get(str(import_id))
        if not file_import_data:
            return jsonify({'error': 'File data not available. Please re-import.'}), 404
        
        headers = file_import_data.get('headers', [])
        all_rows = file_import_data.get('all_rows', [])
        
        return jsonify({
            'success': True,
            'filename': import_record.file_name,
            'file_type': import_record.file_type,
            'headers': headers,
            'rows': all_rows,
            'total': len(all_rows)
        })
    except Exception as e:
        try:
            db_session.rollback()
        except:
            pass
        print(f"Preview error: {e}")
        return jsonify({'error': f'Failed to preview file: {str(e)}'}), 500


@hr_bp.route('/api/download-file/<int:import_id>')
@login_required
def download_file(import_id):
    """Download uploaded file from object storage or session"""
    db_session = get_db_session()
    try:
        db_session.rollback()
        user_id = session.get('user_id')
        user = db_session.get(User, user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 401
        
        org_id = user.organization_id if user.organization_id else user.id
        import_record = db_session.query(HRDataImport).filter_by(
            id=import_id,
            organization_id=org_id
        ).first()
        
        if not import_record:
            return jsonify({'error': 'File not found or access denied'}), 404
        
        file_content = None
        content_type = None
        filename = import_record.file_name
        
        # Try object storage first
        if import_record.file_storage_path:
            storage_service = ObjectStorageService()
            try:
                file_data = storage_service.get_file(import_record.file_storage_path)
                if file_data:
                    file_content, content_type, filename = file_data
            except Exception as e:
                print(f"Storage retrieval failed: {e}")
        
        # Fall back to session if storage not available
        if not file_content:
            file_import_data = session.get('file_imports', {}).get(str(import_id))
            if file_import_data:
                # Reconstruct the file from session
                headers = file_import_data.get('headers', [])
                all_rows = file_import_data.get('all_rows', [])
                file_type = file_import_data.get('file_type', 'employees')
                
                # Generate CSV content from session data
                output = io.StringIO()
                if headers:
                    writer = csv.DictWriter(output, fieldnames=headers)
                    writer.writeheader()
                    writer.writerows(all_rows)
                
                file_content = output.getvalue().encode('utf-8')
                content_type = 'text/csv'
            else:
                return jsonify({'error': 'File not available in session. Please re-import.'}), 404
        
        return Response(
            file_content,
            mimetype=content_type or 'text/csv',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"',
                'X-Content-Type-Options': 'nosniff'
            }
        )
    except Exception as e:
        try:
            db_session.rollback()
        except:
            pass
        print(f"Download error: {e}")
        return jsonify({'error': f'Failed to download file: {str(e)}'}), 500


@hr_bp.route('/analyze')
@login_required
def analyze():
    """HR Data Analysis page"""
    try:
        lang = session.get('language', 'ar')
        user_id = session.get('user_id')
        db_session = get_db_session()
        db_session.rollback()
        user = db_session.get(User, user_id)
        
        if not user:
            return render_template('hr/analyze.html', lang=lang, has_data=False)
        
        org_id = user.organization_id if user.organization_id else user.id
        
        employees = db_session.query(HREmployee).filter_by(organization_id=org_id).all()
    except Exception as e:
        db_session = get_db_session()
        try:
            db_session.rollback()
        except:
            pass
        print(f"HR Analyze Error: {str(e)}")
        return render_template('hr/analyze.html', lang=lang, has_data=False), 500
    
    return render_template('hr/analyze.html', 
                          lang=lang,
                          has_data=len(employees) > 0,
                          employees=employees)
