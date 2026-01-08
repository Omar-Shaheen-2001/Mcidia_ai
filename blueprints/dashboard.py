from flask import Blueprint, render_template, session, current_app, redirect, url_for, jsonify, request, flash, abort, make_response
from flask_jwt_extended import get_jwt_identity
from utils.decorators import login_required
from models import User, Project, AILog, Transaction, Service, ServiceOffering, ChatSession
try:
    from weasyprint import HTML
    WEASYPRINT_AVAILABLE = True
except (ImportError, OSError):
    WEASYPRINT_AVAILABLE = False
    HTML = None
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import json

dashboard_bp = Blueprint('dashboard', __name__)

@dashboard_bp.route('/')
@login_required
def index():
    db = current_app.extensions['sqlalchemy']
    user_id = int(get_jwt_identity())
    user = db.session.query(User).get(user_id)
    
    # Get statistics
    total_projects = db.session.query(Project).filter_by(user_id=user_id).count()
    active_projects = db.session.query(Project).filter_by(user_id=user_id, status='draft').count()
    ai_credits = user.ai_credits_used if user else 0
    total_consultations = db.session.query(ChatSession).filter_by(user_id=user_id).count()
    
    # Get recent projects
    recent_projects = db.session.query(Project).filter_by(user_id=user_id).order_by(Project.updated_at.desc()).limit(5).all()
    
    # Get recent AI logs
    recent_ai_activity = db.session.query(AILog).filter_by(user_id=user_id).order_by(AILog.created_at.desc()).limit(5).all()
    
    # Build service info map for projects
    project_services = {}
    for project in recent_projects:
        if project.module and '_' in project.module:
            parts = project.module.split('_', 1)
            if len(parts) == 2:
                service_slug, offering_slug = parts
                service = db.session.query(Service).filter_by(slug=service_slug).first()
                if service:
                    offering = db.session.query(ServiceOffering).filter_by(
                        service_id=service.id,
                        slug=offering_slug
                    ).first()
                    if offering:
                        project_services[project.id] = {
                            'service': service,
                            'offering': offering
                        }
    
    # Get all active services for dashboard display
    all_services = db.session.query(Service).filter_by(is_active=True).order_by(Service.display_order).limit(8).all()
    
    lang = session.get('language', 'ar')
    
    return render_template('dashboard/index.html', 
                         user=user,
                         total_projects=total_projects,
                         active_projects=active_projects,
                         ai_credits=ai_credits,
                         total_consultations=total_consultations,
                         recent_projects=recent_projects,
                         recent_ai_activity=recent_ai_activity,
                         project_services=project_services,
                         all_services=all_services,
                         lang=lang)

@dashboard_bp.route('/projects')
@login_required
def all_projects():
    """View all projects with pagination and filters"""
    db = current_app.extensions['sqlalchemy']
    user_id = int(get_jwt_identity())
    user = db.session.query(User).get(user_id)
    
    # Pagination
    page = session.get('projects_page', 1) if isinstance(session.get('projects_page'), int) else 1
    per_page = 10
    
    # Get all projects with pagination
    projects_query = db.session.query(Project).filter_by(user_id=user_id).order_by(Project.updated_at.desc())
    projects_paginated = projects_query.paginate(page=page, per_page=per_page, error_out=False)
    
    # Build service info map for all projects
    project_services = {}
    for project in projects_paginated.items:
        if project.module and '_' in project.module:
            parts = project.module.split('_', 1)
            if len(parts) == 2:
                service_slug, offering_slug = parts
                service = db.session.query(Service).filter_by(slug=service_slug).first()
                if service:
                    offering = db.session.query(ServiceOffering).filter_by(
                        service_id=service.id,
                        slug=offering_slug
                    ).first()
                    if offering:
                        project_services[project.id] = {
                            'service': service,
                            'offering': offering
                        }
    
    # Get statistics
    total_projects = projects_query.count()
    draft_projects = db.session.query(Project).filter_by(user_id=user_id, status='draft').count()
    completed_projects = db.session.query(Project).filter_by(user_id=user_id, status='completed').count()
    archived_projects = db.session.query(Project).filter_by(user_id=user_id, status='archived').count()
    
    lang = session.get('language', 'ar')
    
    return render_template('dashboard/projects.html',
                         user=user,
                         projects=projects_paginated.items,
                         pagination=projects_paginated,
                         project_services=project_services,
                         total_projects=total_projects,
                         draft_projects=draft_projects,
                         completed_projects=completed_projects,
                         archived_projects=archived_projects,
                         current_page=page,
                         lang=lang)

@dashboard_bp.route('/project/<int:project_id>')
@login_required
def view_project(project_id):
    """View project details"""
    db = current_app.extensions['sqlalchemy']
    user_id = int(get_jwt_identity())
    
    project = db.session.query(Project).filter_by(id=project_id, user_id=user_id).first_or_404()
    lang = session.get('language', 'ar')
    
    # Get service info
    project_services = {}
    if project.module and '_' in project.module:
        parts = project.module.split('_', 1)
        if len(parts) == 2:
            service_slug, offering_slug = parts
            service = db.session.query(Service).filter_by(slug=service_slug).first()
            if service:
                offering = db.session.query(ServiceOffering).filter_by(
                    service_id=service.id,
                    slug=offering_slug
                ).first()
                if offering:
                    project_services = {'service': service, 'offering': offering}
    
    try:
        content = json.loads(project.content) if project.content else {}
    except:
        content = {}
    
    return render_template('dashboard/view_project.html',
                         project=project,
                         project_services=project_services,
                         content=content,
                         lang=lang)

@dashboard_bp.route('/project/<int:project_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_project(project_id):
    """Edit project content and metadata"""
    db = current_app.extensions['sqlalchemy']
    user_id = int(get_jwt_identity())
    lang = session.get('language', 'ar')
    
    project = db.session.query(Project).filter_by(id=project_id, user_id=user_id).first_or_404()
    
    if request.method == 'POST':
        # Update project content
        new_title = request.form.get('title', project.title)
        new_content = request.form.get('content', project.content)
        
        try:
            project.title = new_title
            project.content = new_content
            db.session.commit()
            flash('تم تحديث المشروع بنجاح / Project updated successfully' if lang == 'ar' else 'Project updated successfully', 'success')
            return redirect(url_for('dashboard.view_project', project_id=project_id))
        except Exception as e:
            db.session.rollback()
            flash(f'خطأ: {str(e)} / Error: {str(e)}' if lang == 'ar' else f'Error: {str(e)}', 'danger')
    
    # Get service info
    project_services = {}
    if project.module and '_' in project.module:
        parts = project.module.split('_', 1)
        if len(parts) == 2:
            service_slug, offering_slug = parts
            service = db.session.query(Service).filter_by(slug=service_slug).first()
            if service:
                offering = db.session.query(ServiceOffering).filter_by(
                    service_id=service.id,
                    slug=offering_slug
                ).first()
                if offering:
                    project_services = {'service': service, 'offering': offering}
    
    return render_template('dashboard/edit_project.html',
                         project=project,
                         project_services=project_services,
                         lang=lang)

@dashboard_bp.route('/api/project/<int:project_id>/delete', methods=['DELETE'])
@login_required
def delete_project(project_id):
    """Delete a project"""
    db = current_app.extensions['sqlalchemy']
    user_id = int(get_jwt_identity())
    
    project = db.session.query(Project).filter_by(id=project_id, user_id=user_id).first()
    if not project:
        return jsonify({'error': 'Project not found'}), 404
    
    try:
        db.session.delete(project)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Project deleted successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@dashboard_bp.route('/api/project/<int:project_id>/export-pdf')
@login_required
def export_project_pdf(project_id):
    """تصدير المشروع كملف PDF مع التنسيق الكامل"""
    db = current_app.extensions['sqlalchemy']
    lang = session.get('language', 'ar')
    
    # Get project
    project = db.session.get(Project, project_id)
    if not project:
        abort(404)
    
    # Check ownership - get JWT identity and handle None
    user_id = get_jwt_identity()
    if user_id is None:
        abort(401)  # Unauthorized
    
    if project.user_id != int(user_id):
        abort(403)  # Forbidden
    
    # Parse project content
    project_data = {}
    try:
        project_data = json.loads(project.content) if project.content else {}
    except:
        project_data = {'input': {}, 'output': ''}
    
    # Get service and offering info
    service = None
    offering = None
    if project.module and '_' in project.module:
        parts = project.module.split('_', 1)
        if len(parts) == 2:
            service_slug, offering_slug = parts
            service = db.session.query(Service).filter_by(slug=service_slug).first()
            if service:
                offering = db.session.query(ServiceOffering).filter_by(
                    service_id=service.id,
                    slug=offering_slug
                ).first()
    
    try:
        # Generate HTML for PDF
        html_content = _generate_project_html(project, service, offering, project_data, lang)
        
        # Convert to PDF using WeasyPrint
        if WEASYPRINT_AVAILABLE:
            try:
                pdf_file = HTML(string=html_content).write_pdf()
                
                # Create response
                response = make_response(pdf_file)
                response.headers['Content-Type'] = 'application/pdf'
                filename = f"consultation_{project_id}.pdf"
                response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
                
                return response
            except Exception as e:
                current_app.logger.error(f"PDF generation error: {str(e)}")
                flash("حدث خطأ أثناء إنشاء ملف PDF (نقص في مكتبات النظام).", "danger")
                return redirect(url_for('dashboard.project_view', project_id=project_id))
        else:
            flash("تصدير PDF غير متاح حالياً على هذا الخادم (نقص في مكتبات النظام).", "warning")
            return redirect(url_for('dashboard.project_view', project_id=project_id))
        
    except Exception as e:
        current_app.logger.error(f"PDF export error: {str(e)}")
        abort(500)

@dashboard_bp.route('/api/project/<int:project_id>/export-excel')
@login_required
def export_project_excel(project_id):
    """تصدير المشروع كملف Excel"""
    db = current_app.extensions['sqlalchemy']
    lang = session.get('language', 'ar')
    
    # Get project
    project = db.session.get(Project, project_id)
    if not project:
        abort(404)
    
    # Check ownership - get JWT identity and handle None
    user_id = get_jwt_identity()
    if user_id is None:
        abort(401)  # Unauthorized
    
    if project.user_id != int(user_id):
        abort(403)  # Forbidden
    
    # Parse project content
    project_data = {}
    try:
        project_data = json.loads(project.content) if project.content else {}
    except:
        project_data = {'input': {}, 'output': ''}
    
    # Get service and offering info
    service = None
    offering = None
    if project.module and '_' in project.module:
        parts = project.module.split('_', 1)
        if len(parts) == 2:
            service_slug, offering_slug = parts
            service = db.session.query(Service).filter_by(slug=service_slug).first()
            if service:
                offering = db.session.query(ServiceOffering).filter_by(
                    service_id=service.id,
                    slug=offering_slug
                ).first()
    
    try:
        from utils.markdown_formatter import extract_sections_from_markdown, clean_markdown_for_excel
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "الاستشارة" if lang == 'ar' else "Consultation"
        ws.sheet_view.rightToLeft = (lang == 'ar')
        
        # Get service color
        service_color = service.color.replace('#', '') if service and service.color else '1e3a8a'
        
        # Define professional styles
        title_fill = PatternFill(start_color=service_color, end_color=service_color, fill_type="solid")
        title_font = Font(name='Arial', size=18, bold=True, color="FFFFFF")
        
        header_fill = PatternFill(start_color=service_color, end_color=service_color, fill_type="solid")
        header_font = Font(name='Arial', size=14, bold=True, color="FFFFFF")
        
        section_fill = PatternFill(start_color="e2e8f0", end_color="e2e8f0", fill_type="solid")
        section_font = Font(name='Arial', size=12, bold=True, color="2c5282")
        
        key_fill = PatternFill(start_color="f7fafc", end_color="f7fafc", fill_type="solid")
        key_font = Font(name='Arial', size=10, bold=True, color="2c5282")
        value_font = Font(name='Arial', size=10)
        
        normal_font = Font(name='Arial', size=10)
        rtl_alignment = Alignment(horizontal='right', vertical='top', wrap_text=True, readingOrder=2)
        ltr_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        row = 1
        
        # Title
        ws[f'A{row}'] = project.title
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].fill = title_fill
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'A{row}:E{row}')
        ws.row_dimensions[row].height = 35
        row += 1
        
        # Service info
        if service and offering:
            ws[f'A{row}'] = f"{service.title_ar if lang == 'ar' else service.title_en} - {offering.title_ar if lang == 'ar' else offering.title_en}"
            ws[f'A{row}'].font = Font(name='Arial', size=11, color="4a5568")
            ws[f'A{row}'].alignment = Alignment(horizontal='center')
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
        
        # Date
        ws[f'A{row}'] = f"{'التاريخ' if lang == 'ar' else 'Date'}: {project.created_at.strftime('%Y-%m-%d %H:%M')}"
        ws[f'A{row}'].font = Font(name='Arial', size=10, color="718096")
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        ws.merge_cells(f'A{row}:E{row}')
        row += 2
        
        # Input section header
        ws[f'A{row}'] = "البيانات المدخلة" if lang == 'ar' else "Input Data"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'A{row}:E{row}')
        ws.row_dimensions[row].height = 30
        row += 1
        
        # Input data
        input_data = project_data.get('input', {})
        for key, value in input_data.items():
            ws[f'A{row}'] = str(key).replace('_', ' ').title()
            ws[f'A{row}'].font = key_font
            ws[f'A{row}'].fill = key_fill
            ws[f'A{row}'].alignment = rtl_alignment if lang == 'ar' else ltr_alignment
            
            ws[f'B{row}'] = str(value)
            ws[f'B{row}'].font = value_font
            ws[f'B{row}'].alignment = rtl_alignment if lang == 'ar' else ltr_alignment
            ws.merge_cells(f'B{row}:E{row}')
            ws.row_dimensions[row].height = 20
            row += 1
        
        row += 1
        
        # Output section header
        ws[f'A{row}'] = "نتيجة الاستشارة الذكية" if lang == 'ar' else "AI Consultation Result"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'A{row}:E{row}')
        ws.row_dimensions[row].height = 30
        row += 1
        
        # Output content with sections
        output_text = project_data.get('output', '')
        sections = extract_sections_from_markdown(output_text)
        
        if sections:
            for section in sections:
                # Section title
                ws[f'A{row}'] = section['title']
                ws[f'A{row}'].font = section_font
                ws[f'A{row}'].fill = section_fill
                ws[f'A{row}'].alignment = rtl_alignment if lang == 'ar' else ltr_alignment
                ws.merge_cells(f'A{row}:E{row}')
                ws.row_dimensions[row].height = 25
                row += 1
                
                # Section content
                content = clean_markdown_for_excel(section['content'])
                ws[f'A{row}'] = content
                ws[f'A{row}'].font = normal_font
                ws[f'A{row}'].alignment = rtl_alignment if lang == 'ar' else ltr_alignment
                ws.merge_cells(f'A{row}:E{row}')
                
                lines = content.count('\n') + 1
                ws.row_dimensions[row].height = max(20, min(lines * 15, 200))
                row += 2
        else:
            content = clean_markdown_for_excel(output_text)
            ws[f'A{row}'] = content
            ws[f'A{row}'].font = normal_font
            ws[f'A{row}'].alignment = rtl_alignment if lang == 'ar' else ltr_alignment
            ws.merge_cells(f'A{row}:E{row}')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        filename = f"consultation_{project_id}.xlsx"
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        current_app.logger.error(f"Excel export error: {str(e)}")
        abort(500)

def _generate_project_html(project, service, offering, project_data, lang):
    """Generate HTML for PDF export with professionally formatted consultation content"""
    from utils.markdown_formatter import format_consultation_output
    
    input_data = project_data.get('input', {})
    output_text = project_data.get('output', '')
    
    # Format Markdown to beautiful HTML with cards, grids, tables, stat boxes
    formatted_output = format_consultation_output(output_text, lang)
    
    # Render the professional export template
    html = render_template(
        'exports/project_export.html',
        project=project,
        service=service,
        offering=offering,
        input_data=input_data,
        formatted_output=formatted_output,
        lang=lang
    )
    
    return html
