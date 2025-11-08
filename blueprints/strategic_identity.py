"""
Strategic Identity Development Module
AI-powered comprehensive organizational identity building
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify, current_app
from flask_jwt_extended import get_jwt_identity
from utils.decorators import login_required
from models import StrategicIdentityProject, StrategicObjective, IdentityKPI, IdentityInitiative
from utils.ai_providers.ai_manager import AIManager
from werkzeug.utils import secure_filename
import json
import os
from datetime import datetime

strategic_identity_bp = Blueprint('strategic_identity', __name__)

UPLOAD_FOLDER = 'static/uploads/strategic_identity'
ALLOWED_EXTENSIONS = {'pdf', 'docx', 'doc', 'xlsx', 'xls', 'txt'}

def get_db():
    """Get database instance"""
    return current_app.extensions['sqlalchemy']

def get_lang():
    """Get current language from session"""
    return session.get('language', 'ar')

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ==================== DASHBOARD ====================

@strategic_identity_bp.route('/')
@login_required
def index():
    """Strategic Identity dashboard - list all projects"""
    db = get_db()
    lang = get_lang()
    user_id = int(get_jwt_identity())
    
    # Get all projects for current user
    projects = db.session.query(StrategicIdentityProject).filter_by(user_id=user_id).order_by(StrategicIdentityProject.updated_at.desc()).all()
    
    return render_template('strategic_identity/index.html', 
                         projects=projects, 
                         lang=lang)

# ==================== CREATE NEW PROJECT ====================

@strategic_identity_bp.route('/create', methods=['GET', 'POST'])
@login_required
def create_project():
    """Create new strategic identity project"""
    db = get_db()
    lang = get_lang()
    user_id = int(get_jwt_identity())
    
    if request.method == 'GET':
        return render_template('strategic_identity/create.html', lang=lang)
    
    try:
        # Extract form data
        organization_name = request.form.get('organization_name')
        sector = request.form.get('sector')
        employee_count = int(request.form.get('employee_count', 0))
        location = request.form.get('location')
        description = request.form.get('description')
        current_objectives = request.form.get('current_objectives')
        ongoing_initiatives = request.form.get('ongoing_initiatives')
        
        # Handle file uploads
        uploaded_files = []
        if 'files' in request.files:
            files = request.files.getlist('files')
            os.makedirs(UPLOAD_FOLDER, exist_ok=True)
            
            for file in files:
                if file and file.filename and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    filename = f"{timestamp}_{filename}"
                    filepath = os.path.join(UPLOAD_FOLDER, filename)
                    file.save(filepath)
                    uploaded_files.append(filepath)
        
        # Create project record
        project = StrategicIdentityProject(
            user_id=user_id,
            organization_name=organization_name,
            sector=sector,
            employee_count=employee_count,
            location=location,
            description=description,
            current_objectives=current_objectives,
            ongoing_initiatives=ongoing_initiatives,
            uploaded_files=json.dumps(uploaded_files) if uploaded_files else None,
            status='draft'
        )
        
        db.session.add(project)
        db.session.commit()
        
        flash(f'تم إنشاء المشروع "{organization_name}" بنجاح / Project created successfully' if lang == 'ar' else f'Project "{organization_name}" created successfully', 'success')
        
        # Redirect to analysis generation
        return redirect(url_for('strategic_identity.generate_analysis', project_id=project.id))
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error creating project: {str(e)}")
        flash(f'خطأ في إنشاء المشروع: {str(e)} / Error creating project: {str(e)}' if lang == 'ar' else f'Error creating project: {str(e)}', 'danger')
        return redirect(url_for('strategic_identity.create_project'))

# ==================== AI ANALYSIS GENERATION ====================

@strategic_identity_bp.route('/project/<int:project_id>/generate-analysis', methods=['GET', 'POST'])
@login_required
def generate_analysis(project_id):
    """Generate AI-powered strategic analysis"""
    db = get_db()
    lang = get_lang()
    user_id = int(get_jwt_identity())
    
    # Authorization check
    project = db.session.query(StrategicIdentityProject).filter_by(id=project_id).first_or_404()
    if project.user_id != user_id:
        flash('غير مصرح لك بالوصول لهذا المشروع / Unauthorized access', 'danger')
        return redirect(url_for('strategic_identity.index'))
    
    if request.method == 'GET':
        return render_template('strategic_identity/generate_analysis.html', 
                             project=project,
                             lang=lang)
    
    try:
        # Create AI provider for strategic analysis
        ai = AIManager.for_use_case('strategic_analysis')
        
        # Build context for AI
        context = f"""
منظمة: {project.organization_name}
القطاع: {project.sector}
عدد الموظفين: {project.employee_count}
الموقع: {project.location}
الوصف: {project.description}
الأهداف الحالية: {project.current_objectives}
المبادرات الجارية: {project.ongoing_initiatives}
"""
        
        # Generate SWOT Analysis
        swot_prompt = f"""أنت خبير استراتيجي. قم بتحليل SWOT شامل للمنظمة التالية:

{context}

قدم تحليل SWOT بصيغة JSON كالتالي:
{{
    "strengths": ["نقطة قوة 1", "نقطة قوة 2", ...],
    "weaknesses": ["نقطة ضعف 1", "نقطة ضعف 2", ...],
    "opportunities": ["فرصة 1", "فرصة 2", ...],
    "threats": ["تهديد 1", "تهديد 2", ...]
}}

قدم على الأقل 5 نقاط لكل عنصر."""

        swot_response = ai.chat(
            prompt=swot_prompt,
            temperature=0.7,
            max_tokens=2000
        )
        
        # Generate Vision & Mission
        identity_prompt = f"""أنت خبير في بناء الهوية الاستراتيجية. بناءً على المعلومات التالية:

{context}

أنشئ:
1. رؤية ملهمة (Vision) - جملة واحدة طموحة
2. رسالة واضحة (Mission) - فقرة مختصرة تصف الهدف الأساسي
3. 5 قيم جوهرية (Core Values) مع وصف مختصر لكل قيمة
4. 4 مجالات استراتيجية رئيسية (Strategic Themes)

قدم النتيجة بصيغة JSON:
{{
    "vision": "الرؤية هنا...",
    "mission": "الرسالة هنا...",
    "core_values": [
        {{"value": "النزاهة", "description": "وصف مختصر"}},
        ...
    ],
    "strategic_themes": ["مجال 1", "مجال 2", ...]
}}"""

        identity_response = ai.chat(
            prompt=identity_prompt,
            temperature=0.7,
            max_tokens=2000
        )
        
        # Parse responses
        try:
            swot_data = json.loads(swot_response)
            identity_data = json.loads(identity_response)
        except json.JSONDecodeError:
            # Fallback: try to extract JSON from response
            import re
            swot_match = re.search(r'\{.*\}', swot_response, re.DOTALL)
            identity_match = re.search(r'\{.*\}', identity_response, re.DOTALL)
            
            swot_data = json.loads(swot_match.group(0)) if swot_match else {}
            identity_data = json.loads(identity_match.group(0)) if identity_match else {}
        
        # Update project with AI-generated data
        project.swot_analysis = json.dumps(swot_data, ensure_ascii=False)
        project.vision_statement = identity_data.get('vision', '')
        project.mission_statement = identity_data.get('mission', '')
        project.core_values = json.dumps(identity_data.get('core_values', []), ensure_ascii=False)
        project.strategic_themes = json.dumps(identity_data.get('strategic_themes', []), ensure_ascii=False)
        project.status = 'analysis_complete'
        project.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        flash('تم إنشاء التحليل الاستراتيجي بنجاح / Strategic analysis generated successfully', 'success')
        return redirect(url_for('strategic_identity.project_dashboard', project_id=project.id))
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error generating analysis: {str(e)}")
        flash(f'خطأ في إنشاء التحليل: {str(e)} / Error generating analysis: {str(e)}', 'danger')
        return redirect(url_for('strategic_identity.generate_analysis', project_id=project_id))

# ==================== PROJECT DASHBOARD ====================

@strategic_identity_bp.route('/project/<int:project_id>')
@login_required
def project_dashboard(project_id):
    """View strategic identity project dashboard"""
    db = get_db()
    lang = get_lang()
    user_id = int(get_jwt_identity())
    
    # Authorization check
    project = db.session.query(StrategicIdentityProject).filter_by(id=project_id).first_or_404()
    if project.user_id != user_id:
        flash('غير مصرح لك بالوصول لهذا المشروع / Unauthorized access', 'danger')
        return redirect(url_for('strategic_identity.index'))
    
    # Parse JSON data
    swot = json.loads(project.swot_analysis) if project.swot_analysis else {}
    pestel = json.loads(project.pestel_analysis) if project.pestel_analysis else {}
    values = json.loads(project.core_values) if project.core_values else []
    themes = json.loads(project.strategic_themes) if project.strategic_themes else []
    
    # Get related data
    objectives = db.session.query(StrategicObjective).filter_by(project_id=project_id).all()
    kpis = db.session.query(IdentityKPI).filter_by(project_id=project_id).all()
    initiatives = db.session.query(IdentityInitiative).filter_by(project_id=project_id).all()
    
    return render_template('strategic_identity/dashboard.html',
                         project=project,
                         swot=swot,
                         pestel=pestel,
                         values=values,
                         themes=themes,
                         objectives=objectives,
                         kpis=kpis,
                         initiatives=initiatives,
                         lang=lang)

# ==================== EXPORT FUNCTIONS ====================

@strategic_identity_bp.route('/project/<int:project_id>/export/pdf')
@login_required
def export_pdf(project_id):
    """Export strategic identity to PDF"""
    from flask import make_response
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER
    from io import BytesIO
    import arabic_reshaper
    from bidi.algorithm import get_display
    
    db = get_db()
    lang = get_lang()
    user_id = int(get_jwt_identity())
    
    # Authorization check
    project = db.session.query(StrategicIdentityProject).filter_by(id=project_id).first_or_404()
    if project.user_id != user_id:
        flash('غير مصرح / Unauthorized', 'danger')
        return redirect(url_for('strategic_identity.index'))
    
    # Parse JSON data
    swot = json.loads(project.swot_analysis) if project.swot_analysis else {}
    pestel = json.loads(project.pestel_analysis) if project.pestel_analysis else {}
    values = json.loads(project.core_values) if project.core_values else []
    themes = json.loads(project.strategic_themes) if project.strategic_themes else []
    
    # Get related data
    objectives = db.session.query(StrategicObjective).filter_by(project_id=project_id).all()
    kpis = db.session.query(IdentityKPI).filter_by(project_id=project_id).all()
    initiatives = db.session.query(IdentityInitiative).filter_by(project_id=project_id).all()
    
    def prep_text(text):
        """Prepare Arabic text for PDF"""
        if not text:
            return ""
        reshaped = arabic_reshaper.reshape(str(text))
        return get_display(reshaped)
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Arabic RTL style
    rtl_style = ParagraphStyle(
        'RTL',
        parent=styles['Normal'],
        alignment=TA_RIGHT,
        fontSize=12,
        leading=16
    )
    
    title_style = ParagraphStyle(
        'RTLTitle',
        parent=styles['Heading1'],
        alignment=TA_CENTER,
        fontSize=18,
        leading=22
    )
    
    heading_style = ParagraphStyle(
        'RTLHeading',
        parent=styles['Heading2'],
        alignment=TA_RIGHT,
        fontSize=14,
        leading=18,
        spaceAfter=10
    )
    
    # Title
    title_text = prep_text(f"الهوية الاستراتيجية - {project.organization_name}")
    elements.append(Paragraph(title_text, title_style))
    elements.append(Spacer(1, 0.5*cm))
    
    # Vision
    if project.vision_statement:
        elements.append(Paragraph(prep_text("الرؤية:"), heading_style))
        elements.append(Paragraph(prep_text(project.vision_statement), rtl_style))
        elements.append(Spacer(1, 0.3*cm))
    
    # Mission
    if project.mission_statement:
        elements.append(Paragraph(prep_text("الرسالة:"), heading_style))
        elements.append(Paragraph(prep_text(project.mission_statement), rtl_style))
        elements.append(Spacer(1, 0.5*cm))
    
    # Core Values
    if values:
        elements.append(Paragraph(prep_text("القيم المؤسسية:"), heading_style))
        for value in values:
            # Handle both dict and string formats
            if isinstance(value, dict):
                val_text = f"• {value.get('value', '')} - {value.get('description', '')}"
            else:
                val_text = f"• {str(value)}"
            elements.append(Paragraph(prep_text(val_text), rtl_style))
        elements.append(Spacer(1, 0.5*cm))
    
    # Strategic Themes
    if themes:
        elements.append(Paragraph(prep_text("المجالات الاستراتيجية:"), heading_style))
        for theme in themes:
            # Handle both dict and string formats
            if isinstance(theme, dict):
                theme_text = f"• {theme.get('theme', '')} - {theme.get('description', '')}"
            else:
                theme_text = f"• {str(theme)}"
            elements.append(Paragraph(prep_text(theme_text), rtl_style))
        elements.append(Spacer(1, 0.5*cm))
    
    elements.append(PageBreak())
    
    # SWOT Analysis
    if swot:
        elements.append(Paragraph(prep_text("تحليل SWOT:"), heading_style))
        elements.append(Spacer(1, 0.2*cm))
        
        swot_data = []
        swot_data.append([prep_text("نقاط القوة"), prep_text("نقاط الضعف")])
        
        strengths = swot.get('strengths', [])
        weaknesses = swot.get('weaknesses', [])
        max_len = max(len(strengths), len(weaknesses))
        
        for i in range(max_len):
            s = prep_text(strengths[i]) if i < len(strengths) else ""
            w = prep_text(weaknesses[i]) if i < len(weaknesses) else ""
            swot_data.append([s, w])
        
        swot_table = Table(swot_data, colWidths=[8*cm, 8*cm])
        swot_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(swot_table)
        elements.append(Spacer(1, 0.5*cm))
        
        # Opportunities and Threats
        swot_data2 = []
        swot_data2.append([prep_text("الفرص"), prep_text("التهديدات")])
        
        opportunities = swot.get('opportunities', [])
        threats = swot.get('threats', [])
        max_len2 = max(len(opportunities), len(threats))
        
        for i in range(max_len2):
            o = prep_text(opportunities[i]) if i < len(opportunities) else ""
            t = prep_text(threats[i]) if i < len(threats) else ""
            swot_data2.append([o, t])
        
        swot_table2 = Table(swot_data2, colWidths=[8*cm, 8*cm])
        swot_table2.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(swot_table2)
    
    # PESTEL Analysis
    if pestel:
        elements.append(PageBreak())
        elements.append(Paragraph(prep_text("تحليل PESTEL:"), heading_style))
        elements.append(Spacer(1, 0.2*cm))
        
        pestel_categories = [
            ('political', 'العوامل السياسية'),
            ('economic', 'العوامل الاقتصادية'),
            ('social', 'العوامل الاجتماعية'),
            ('technological', 'العوامل التقنية'),
            ('environmental', 'العوامل البيئية'),
            ('legal', 'العوامل القانونية')
        ]
        
        for key, label in pestel_categories:
            if key in pestel and pestel[key]:
                elements.append(Paragraph(prep_text(f"{label}:"), rtl_style))
                for item in pestel[key]:
                    elements.append(Paragraph(prep_text(f"• {item}"), rtl_style))
                elements.append(Spacer(1, 0.3*cm))
    
    # Strategic Objectives
    if objectives:
        elements.append(PageBreak())
        elements.append(Paragraph(prep_text("الأهداف الاستراتيجية:"), heading_style))
        elements.append(Spacer(1, 0.2*cm))
        
        obj_data = [[prep_text("الهدف"), prep_text("الوصف"), prep_text("الإطار الزمني")]]
        for obj in objectives:
            obj_data.append([
                prep_text(obj.objective_name),
                prep_text(obj.description or ''),
                prep_text(obj.timeframe or '')
            ])
        
        obj_table = Table(obj_data, colWidths=[5*cm, 7*cm, 4*cm])
        obj_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(obj_table)
    
    # KPIs
    if kpis:
        elements.append(PageBreak())
        elements.append(Paragraph(prep_text("مؤشرات الأداء الرئيسية:"), heading_style))
        elements.append(Spacer(1, 0.2*cm))
        
        kpi_data = [[prep_text("المؤشر"), prep_text("القيمة المستهدفة"), prep_text("الوحدة")]]
        for kpi in kpis:
            kpi_data.append([
                prep_text(kpi.kpi_name),
                prep_text(str(kpi.target_value)),
                prep_text(kpi.measurement_unit or '')
            ])
        
        kpi_table = Table(kpi_data, colWidths=[8*cm, 4*cm, 4*cm])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(kpi_table)
    
    # Initiatives
    if initiatives:
        elements.append(PageBreak())
        elements.append(Paragraph(prep_text("المبادرات الاستراتيجية:"), heading_style))
        elements.append(Spacer(1, 0.2*cm))
        
        init_data = [[prep_text("المبادرة"), prep_text("الوصف"), prep_text("الميزانية")]]
        for init in initiatives:
            budget = f"{init.budget or 0} ر.س" if init.budget else ""
            init_data.append([
                prep_text(init.initiative_name),
                prep_text(init.description or ''),
                prep_text(budget)
            ])
        
        init_table = Table(init_data, colWidths=[5*cm, 7*cm, 4*cm])
        init_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(init_table)
    
    doc.build(elements)
    
    pdf = buffer.getvalue()
    buffer.close()
    
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=strategic_identity_{project_id}.pdf'
    
    return response

@strategic_identity_bp.route('/project/<int:project_id>/export/excel')
@login_required
def export_excel(project_id):
    """Export strategic identity to Excel"""
    from flask import make_response
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    
    db = get_db()
    user_id = int(get_jwt_identity())
    
    # Authorization check
    project = db.session.query(StrategicIdentityProject).filter_by(id=project_id).first_or_404()
    if project.user_id != user_id:
        flash('غير مصرح / Unauthorized', 'danger')
        return redirect(url_for('strategic_identity.index'))
    
    # Parse JSON data
    swot = json.loads(project.swot_analysis) if project.swot_analysis else {}
    pestel = json.loads(project.pestel_analysis) if project.pestel_analysis else {}
    values = json.loads(project.core_values) if project.core_values else []
    themes = json.loads(project.strategic_themes) if project.strategic_themes else []
    
    # Get related data
    objectives = db.session.query(StrategicObjective).filter_by(project_id=project_id).all()
    kpis = db.session.query(IdentityKPI).filter_by(project_id=project_id).all()
    initiatives = db.session.query(IdentityInitiative).filter_by(project_id=project_id).all()
    
    # Create workbook
    wb = Workbook()
    
    # Overview Sheet
    ws_overview = wb.active
    ws_overview.title = "Overview"
    ws_overview.sheet_view.rightToLeft = True
    
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=14)
    subheader_font = Font(bold=True, size=12)
    
    ws_overview['A1'] = "الهوية الاستراتيجية"
    ws_overview['A1'].font = header_font
    ws_overview['A1'].fill = header_fill
    ws_overview.merge_cells('A1:B1')
    
    row = 3
    ws_overview[f'A{row}'] = "اسم المؤسسة"
    ws_overview[f'B{row}'] = project.organization_name
    row += 1
    ws_overview[f'A{row}'] = "القطاع"
    ws_overview[f'B{row}'] = project.sector
    row += 1
    ws_overview[f'A{row}'] = "عدد الموظفين"
    ws_overview[f'B{row}'] = project.employee_count
    row += 2
    
    if project.vision_statement:
        ws_overview[f'A{row}'] = "الرؤية"
        ws_overview[f'A{row}'].font = Font(bold=True)
        row += 1
        ws_overview[f'A{row}'] = project.vision_statement
        ws_overview.merge_cells(f'A{row}:B{row}')
        row += 2
    
    if project.mission_statement:
        ws_overview[f'A{row}'] = "الرسالة"
        ws_overview[f'A{row}'].font = Font(bold=True)
        row += 1
        ws_overview[f'A{row}'] = project.mission_statement
        ws_overview.merge_cells(f'A{row}:B{row}')
        row += 2
    
    ws_overview.column_dimensions['A'].width = 20
    ws_overview.column_dimensions['B'].width = 50
    
    # SWOT Sheet
    if swot:
        ws_swot = wb.create_sheet("SWOT Analysis")
        ws_swot.sheet_view.rightToLeft = True
        
        ws_swot['A1'] = "تحليل SWOT"
        ws_swot['A1'].font = header_font
        ws_swot['A1'].fill = header_fill
        ws_swot.merge_cells('A1:D1')
        
        ws_swot['A3'] = "نقاط القوة"
        ws_swot['B3'] = "نقاط الضعف"
        ws_swot['C3'] = "الفرص"
        ws_swot['D3'] = "التهديدات"
        
        for cell in ['A3', 'B3', 'C3', 'D3']:
            ws_swot[cell].font = Font(bold=True)
            ws_swot[cell].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        row = 4
        max_items = max(
            len(swot.get('strengths', [])),
            len(swot.get('weaknesses', [])),
            len(swot.get('opportunities', [])),
            len(swot.get('threats', []))
        )
        
        for i in range(max_items):
            strengths = swot.get('strengths', [])
            weaknesses = swot.get('weaknesses', [])
            opportunities = swot.get('opportunities', [])
            threats = swot.get('threats', [])
            
            ws_swot[f'A{row}'] = strengths[i] if i < len(strengths) else ""
            ws_swot[f'B{row}'] = weaknesses[i] if i < len(weaknesses) else ""
            ws_swot[f'C{row}'] = opportunities[i] if i < len(opportunities) else ""
            ws_swot[f'D{row}'] = threats[i] if i < len(threats) else ""
            row += 1
        
        for col in ['A', 'B', 'C', 'D']:
            ws_swot.column_dimensions[col].width = 30
    
    # Core Values Sheet
    if values:
        ws_values = wb.create_sheet("Core Values")
        ws_values.sheet_view.rightToLeft = True
        
        ws_values['A1'] = "القيم المؤسسية"
        ws_values['A1'].font = header_font
        ws_values['A1'].fill = header_fill
        ws_values.merge_cells('A1:B1')
        
        ws_values['A3'] = "القيمة"
        ws_values['B3'] = "الوصف"
        ws_values['A3'].font = Font(bold=True)
        ws_values['B3'].font = Font(bold=True)
        
        row = 4
        for value in values:
            # Handle both dict and string formats
            if isinstance(value, dict):
                ws_values[f'A{row}'] = value.get('value', '')
                ws_values[f'B{row}'] = value.get('description', '')
            else:
                ws_values[f'A{row}'] = str(value)
                ws_values[f'B{row}'] = ''
            row += 1
        
        ws_values.column_dimensions['A'].width = 20
        ws_values.column_dimensions['B'].width = 50
    
    # Strategic Themes Sheet
    if themes:
        ws_themes = wb.create_sheet("Strategic Themes")
        ws_themes.sheet_view.rightToLeft = True
        
        ws_themes['A1'] = "المجالات الاستراتيجية"
        ws_themes['A1'].font = header_font
        ws_themes['A1'].fill = header_fill
        ws_themes.merge_cells('A1:B1')
        
        ws_themes['A3'] = "المجال"
        ws_themes['B3'] = "الوصف"
        ws_themes['A3'].font = Font(bold=True)
        ws_themes['B3'].font = Font(bold=True)
        
        row = 4
        for theme in themes:
            # Handle both dict and string formats
            if isinstance(theme, dict):
                ws_themes[f'A{row}'] = theme.get('theme', '')
                ws_themes[f'B{row}'] = theme.get('description', '')
            else:
                ws_themes[f'A{row}'] = str(theme)
                ws_themes[f'B{row}'] = ''
            row += 1
        
        ws_themes.column_dimensions['A'].width = 25
        ws_themes.column_dimensions['B'].width = 50
    
    # PESTEL Analysis Sheet
    if pestel:
        ws_pestel = wb.create_sheet("PESTEL Analysis")
        ws_pestel.sheet_view.rightToLeft = True
        
        ws_pestel['A1'] = "تحليل PESTEL"
        ws_pestel['A1'].font = header_font
        ws_pestel['A1'].fill = header_fill
        ws_pestel.merge_cells('A1:B1')
        
        row = 3
        pestel_categories = [
            ('political', 'العوامل السياسية'),
            ('economic', 'العوامل الاقتصادية'),
            ('social', 'العوامل الاجتماعية'),
            ('technological', 'العوامل التقنية'),
            ('environmental', 'العوامل البيئية'),
            ('legal', 'العوامل القانونية')
        ]
        
        for key, label in pestel_categories:
            if key in pestel and pestel[key]:
                ws_pestel[f'A{row}'] = label
                ws_pestel[f'A{row}'].font = subheader_font
                row += 1
                for item in pestel[key]:
                    ws_pestel[f'A{row}'] = item
                    row += 1
                row += 1
        
        ws_pestel.column_dimensions['A'].width = 60
    
    # Strategic Objectives Sheet
    if objectives:
        ws_obj = wb.create_sheet("Strategic Objectives")
        ws_obj.sheet_view.rightToLeft = True
        
        ws_obj['A1'] = "الأهداف الاستراتيجية"
        ws_obj['A1'].font = header_font
        ws_obj['A1'].fill = header_fill
        ws_obj.merge_cells('A1:C1')
        
        ws_obj['A3'] = "الهدف"
        ws_obj['B3'] = "الوصف"
        ws_obj['C3'] = "الإطار الزمني"
        for cell in ['A3', 'B3', 'C3']:
            ws_obj[cell].font = Font(bold=True)
            ws_obj[cell].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        row = 4
        for obj in objectives:
            ws_obj[f'A{row}'] = obj.objective_name
            ws_obj[f'B{row}'] = obj.description or ''
            ws_obj[f'C{row}'] = obj.timeframe or ''
            row += 1
        
        ws_obj.column_dimensions['A'].width = 25
        ws_obj.column_dimensions['B'].width = 40
        ws_obj.column_dimensions['C'].width = 20
    
    # KPIs Sheet
    if kpis:
        ws_kpi = wb.create_sheet("KPIs")
        ws_kpi.sheet_view.rightToLeft = True
        
        ws_kpi['A1'] = "مؤشرات الأداء الرئيسية"
        ws_kpi['A1'].font = header_font
        ws_kpi['A1'].fill = header_fill
        ws_kpi.merge_cells('A1:D1')
        
        ws_kpi['A3'] = "المؤشر"
        ws_kpi['B3'] = "القيمة المستهدفة"
        ws_kpi['C3'] = "الوحدة"
        ws_kpi['D3'] = "التكرار"
        for cell in ['A3', 'B3', 'C3', 'D3']:
            ws_kpi[cell].font = Font(bold=True)
            ws_kpi[cell].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        row = 4
        for kpi in kpis:
            ws_kpi[f'A{row}'] = kpi.kpi_name
            ws_kpi[f'B{row}'] = kpi.target_value
            ws_kpi[f'C{row}'] = kpi.measurement_unit or ''
            ws_kpi[f'D{row}'] = kpi.frequency or ''
            row += 1
        
        ws_kpi.column_dimensions['A'].width = 30
        ws_kpi.column_dimensions['B'].width = 15
        ws_kpi.column_dimensions['C'].width = 15
        ws_kpi.column_dimensions['D'].width = 15
    
    # Initiatives Sheet
    if initiatives:
        ws_init = wb.create_sheet("Initiatives")
        ws_init.sheet_view.rightToLeft = True
        
        ws_init['A1'] = "المبادرات الاستراتيجية"
        ws_init['A1'].font = header_font
        ws_init['A1'].fill = header_fill
        ws_init.merge_cells('A1:D1')
        
        ws_init['A3'] = "المبادرة"
        ws_init['B3'] = "الوصف"
        ws_init['C3'] = "الميزانية"
        ws_init['D3'] = "المسؤول"
        for cell in ['A3', 'B3', 'C3', 'D3']:
            ws_init[cell].font = Font(bold=True)
            ws_init[cell].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        row = 4
        for init in initiatives:
            ws_init[f'A{row}'] = init.initiative_name
            ws_init[f'B{row}'] = init.description or ''
            ws_init[f'C{row}'] = f"{init.budget or 0} ر.س" if init.budget else ""
            ws_init[f'D{row}'] = init.responsible_party or ''
            row += 1
        
        ws_init.column_dimensions['A'].width = 25
        ws_init.column_dimensions['B'].width = 40
        ws_init.column_dimensions['C'].width = 15
        ws_init.column_dimensions['D'].width = 20
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=strategic_identity_{project_id}.xlsx'
    
    return response
