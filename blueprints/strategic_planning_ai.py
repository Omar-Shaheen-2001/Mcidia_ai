"""
Strategic Planning & KPIs Development Module
AI-powered strategic planning with SWOT, PESTEL, Vision/Mission, Goals, and KPIs generation
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify, current_app, make_response
from flask_jwt_extended import get_jwt_identity
from utils.decorators import login_required
from models import StrategicPlan, StrategicKPI, StrategicInitiative, ServiceOffering, User
from utils.ai_providers.ai_manager import AIManager
import json
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

strategic_planning_bp = Blueprint('strategic_planning_ai', __name__)

def get_db():
    """Get database instance"""
    return current_app.extensions['sqlalchemy']

def get_lang():
    """Get current language from session"""
    return session.get('language', 'ar')

# ==================== DASHBOARD ====================

@strategic_planning_bp.route('/')
@login_required
def index():
    """Strategic Planning dashboard - list all plans"""
    db = get_db()
    lang = get_lang()
    user_id = int(get_jwt_identity())
    
    # Get all plans for current user
    plans = db.session.query(StrategicPlan).filter_by(user_id=user_id).order_by(StrategicPlan.updated_at.desc()).all()
    
    return render_template('strategic_planning/index.html', 
                         plans=plans, 
                         lang=lang)

# ==================== CREATE NEW PLAN ====================

@strategic_planning_bp.route('/create', methods=['GET', 'POST'])
@login_required
def create_plan():
    """Create new strategic plan - step 1: organization information"""
    db = get_db()
    lang = get_lang()
    user_id = int(get_jwt_identity())
    
    if request.method == 'GET':
        # Get offering details for form configuration
        offering = db.session.query(ServiceOffering).filter_by(slug='strategic-planning-kpis').first()
        form_fields = json.loads(offering.form_fields) if offering and offering.form_fields else []
        
        return render_template('strategic_planning/create.html', 
                             form_fields=form_fields,
                             lang=lang)
    
    try:
        # Extract form data
        organization_name = request.form.get('organization_name')
        industry_sector = request.form.get('industry_sector')
        employee_count = int(request.form.get('employee_count', 0))
        planning_period = int(request.form.get('planning_period', 3))
        organization_description = request.form.get('organization_description')
        current_challenges = request.form.get('current_challenges')
        opportunities = request.form.get('opportunities')
        strategic_priorities = request.form.get('strategic_priorities', '')
        
        # Create plan record
        current_year = datetime.now().year
        plan = StrategicPlan(
            user_id=user_id,
            title=f"{organization_name} - {lang == 'ar' and 'الخطة الاستراتيجية' or 'Strategic Plan'} {current_year}-{current_year + planning_period}",
            planning_period=f"{current_year}-{current_year + planning_period}",
            start_year=current_year,
            end_year=current_year + planning_period,
            industry_sector=industry_sector,
            employee_count=employee_count,
            organization_description=organization_description,
            current_challenges=json.dumps([c.strip() for c in (current_challenges or '').split('\n') if c.strip()]),
            opportunities=json.dumps([o.strip() for o in (opportunities or '').split('\n') if o.strip()]),
            status='draft'
        )
        
        db.session.add(plan)
        db.session.commit()
        
        flash('تم إنشاء الخطة بنجاح / Plan created successfully' if lang == 'ar' else 'Plan created successfully', 'success')
        return redirect(url_for('strategic_planning_ai.analyze_swot', plan_id=plan.id))
        
    except Exception as e:
        db.session.rollback()
        flash(f'خطأ: {str(e)} / Error: {str(e)}' if lang == 'ar' else f'Error: {str(e)}', 'danger')
        return redirect(url_for('strategic_planning_ai.create_plan'))

# ==================== SWOT ANALYSIS (AI-POWERED) ====================

@strategic_planning_bp.route('/plan/<int:plan_id>/analyze-swot')
@login_required
def analyze_swot(plan_id):
    """Display SWOT analysis page"""
    db = get_db()
    lang = get_lang()
    user_id = int(get_jwt_identity())
    
    # Authorization: Verify ownership
    plan = db.session.query(StrategicPlan).filter_by(id=plan_id).first_or_404()
    if plan.user_id != user_id:
        flash('غير مصرح لك بالوصول لهذه الخطة / Unauthorized access' if lang == 'ar' else 'Unauthorized access', 'danger')
        return redirect(url_for('strategic_planning_ai.index')), 403
    
    # Parse existing SWOT if available
    swot_data = json.loads(plan.swot_analysis) if plan.swot_analysis else None
    
    return render_template('strategic_planning/swot_analysis.html',
                         plan=plan,
                         swot_data=swot_data,
                         lang=lang)

@strategic_planning_bp.route('/plan/<int:plan_id>/generate-swot', methods=['POST'])
@login_required
def generate_swot(plan_id):
    """Generate SWOT analysis using AI"""
    db = get_db()
    lang = get_lang()
    user_id = int(get_jwt_identity())
    
    # Authorization: Verify ownership
    plan = db.session.query(StrategicPlan).filter_by(id=plan_id).first_or_404()
    if plan.user_id != user_id:
        return jsonify({'success': False, 'error': 'Unauthorized access'}), 403
    
    try:
        # Prepare context for AI
        challenges = json.loads(plan.current_challenges) if plan.current_challenges else []
        opportunities_list = json.loads(plan.opportunities) if plan.opportunities else []
        
        prompt = f"""قم بإجراء تحليل SWOT شامل للمؤسسة التالية:

**اسم المؤسسة:** {plan.title}
**القطاع:** {plan.industry_sector}
**عدد الموظفين:** {plan.employee_count}

**وصف المؤسسة:**
{plan.organization_description}

**التحديات الحالية:**
{chr(10).join(f'- {c}' for c in challenges)}

**الفرص المتاحة:**
{chr(10).join(f'- {o}' for o in opportunities_list)}

يرجى تقديم تحليل SWOT تفصيلي يشمل:
1. نقاط القوة (Strengths): 5-7 نقاط على الأقل
2. نقاط الضعف (Weaknesses): 5-7 نقاط على الأقل
3. الفرص (Opportunities): 5-7 نقاط على الأقل
4. التهديدات (Threats): 5-7 نقاط على الأقل

الرد بصيغة JSON:
{{
  "strengths": ["نقطة قوة 1", "نقطة قوة 2", ...],
  "weaknesses": ["نقطة ضعف 1", "نقطة ضعف 2", ...],
  "opportunities": ["فرصة 1", "فرصة 2", ...],
  "threats": ["تهديد 1", "تهديد 2", ...]
}}"""
        
        # Call AI using HuggingFace (FREE!)
        ai = AIManager.for_use_case('swot_analysis')
        system_prompt = "أنت مستشار استراتيجي خبير متخصص في تحليل SWOT للمؤسسات. قدّم تحليلاً شاملاً ودقيقاً بصيغة JSON."
        response = ai.chat(
            prompt=prompt,
            system_prompt=system_prompt,
            response_format='json'
        )
        
        # Parse response
        swot_data = json.loads(response)
        
        # Save to database
        plan.swot_analysis = json.dumps(swot_data, ensure_ascii=False)
        db.session.commit()
        
        return jsonify({'success': True, 'data': swot_data})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== PESTEL ANALYSIS (AI-POWERED) ====================

@strategic_planning_bp.route('/plan/<int:plan_id>/analyze-pestel')
@login_required
def analyze_pestel(plan_id):
    """Display PESTEL analysis page"""
    db = get_db()
    lang = get_lang()
    user_id = int(get_jwt_identity())
    
    # Authorization: Verify ownership
    plan = db.session.query(StrategicPlan).filter_by(id=plan_id).first_or_404()
    if plan.user_id != user_id:
        flash('غير مصرح لك بالوصول لهذه الخطة / Unauthorized access' if lang == 'ar' else 'Unauthorized access', 'danger')
        return redirect(url_for('strategic_planning_ai.index')), 403
    
    # Parse existing PESTEL if available
    pestel_data = json.loads(plan.pestel_analysis) if plan.pestel_analysis else None
    
    return render_template('strategic_planning/pestel_analysis.html',
                         plan=plan,
                         pestel_data=pestel_data,
                         lang=lang)

@strategic_planning_bp.route('/plan/<int:plan_id>/generate-pestel', methods=['POST'])
@login_required
def generate_pestel(plan_id):
    """Generate PESTEL analysis using AI"""
    db = get_db()
    lang = get_lang()
    user_id = int(get_jwt_identity())
    
    # Authorization: Verify ownership
    plan = db.session.query(StrategicPlan).filter_by(id=plan_id).first_or_404()
    if plan.user_id != user_id:
        return jsonify({'success': False, 'error': 'Unauthorized access'}), 403
    
    try:
        prompt = f"""قم بإجراء تحليل PESTEL شامل للمؤسسة التالية:

**اسم المؤسسة:** {plan.title}
**القطاع:** {plan.industry_sector}
**وصف المؤسسة:** {plan.organization_description}

يرجى تحليل العوامل الخارجية المؤثرة على المؤسسة:
1. السياسية (Political): 3-5 عوامل
2. الاقتصادية (Economic): 3-5 عوامل
3. الاجتماعية (Social): 3-5 عوامل
4. التقنية (Technological): 3-5 عوامل
5. البيئية (Environmental): 3-5 عوامل
6. القانونية (Legal): 3-5 عوامل

الرد بصيغة JSON:
{{
  "political": ["عامل سياسي 1", ...],
  "economic": ["عامل اقتصادي 1", ...],
  "social": ["عامل اجتماعي 1", ...],
  "technological": ["عامل تقني 1", ...],
  "environmental": ["عامل بيئي 1", ...],
  "legal": ["عامل قانوني 1", ...]
}}"""
        
        # Call AI using HuggingFace (FREE!)
        ai = AIManager.for_use_case('pestel_analysis')
        system_prompt = "أنت مستشار استراتيجي خبير متخصص في تحليل PESTEL للبيئة الخارجية للمؤسسات. قدّم تحليلاً شاملاً ودقيقاً بصيغة JSON."
        response = ai.chat(
            prompt=prompt,
            system_prompt=system_prompt,
            response_format='json'
        )
        
        # Parse response
        pestel_data = json.loads(response)
        
        # Save to database
        plan.pestel_analysis = json.dumps(pestel_data, ensure_ascii=False)
        db.session.commit()
        
        return jsonify({'success': True, 'data': pestel_data})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== STRATEGIC FRAMEWORK GENERATION ====================

@strategic_planning_bp.route('/plan/<int:plan_id>/generate-framework')
@login_required
def generate_framework(plan_id):
    """Generate Vision, Mission, Values, and Goals using AI"""
    db = get_db()
    lang = get_lang()
    user_id = int(get_jwt_identity())
    
    # Authorization: Verify ownership
    plan = db.session.query(StrategicPlan).filter_by(id=plan_id).first_or_404()
    if plan.user_id != user_id:
        flash('غير مصرح لك بالوصول لهذه الخطة / Unauthorized access' if lang == 'ar' else 'Unauthorized access', 'danger')
        return redirect(url_for('strategic_planning_ai.index')), 403
    
    return render_template('strategic_planning/framework.html',
                         plan=plan,
                         lang=lang)

@strategic_planning_bp.route('/plan/<int:plan_id>/ai-generate-framework', methods=['POST'])
@login_required
def ai_generate_framework(plan_id):
    """AI generates vision, mission, values, and strategic goals"""
    db = get_db()
    user_id = int(get_jwt_identity())
    
    # Authorization: Verify ownership
    plan = db.session.query(StrategicPlan).filter_by(id=plan_id).first_or_404()
    if plan.user_id != user_id:
        return jsonify({'success': False, 'error': 'Unauthorized access'}), 403
    
    try:
        # Get SWOT data
        swot = json.loads(plan.swot_analysis) if plan.swot_analysis else {}
        
        prompt = f"""بناءً على التحليل الاستراتيجي للمؤسسة التالية، قم بتوليد:

**اسم المؤسسة:** {plan.title}
**القطاع:** {plan.industry_sector}
**الوصف:** {plan.organization_description}

**نقاط القوة الرئيسية:**
{chr(10).join(f'- {s}' for s in swot.get('strengths', [])[:3])}

**الفرص الرئيسية:**
{chr(10).join(f'- {o}' for o in swot.get('opportunities', [])[:3])}

يرجى توليد:
1. **الرؤية (Vision)**: بيان ملهم عن مستقبل المؤسسة (جملة واحدة قوية)
2. **الرسالة (Mission)**: الهدف الأساسي للمؤسسة (فقرة قصيرة)
3. **القيم الجوهرية (Core Values)**: 5-7 قيم أساسية مع شرح موجز لكل قيمة
4. **الأهداف الاستراتيجية (Strategic Goals)**: 4-6 أهداف SMART تغطي المجالات الرئيسية

الرد بصيغة JSON:
{{
  "vision": "بيان الرؤية",
  "mission": "بيان الرسالة",
  "values": [
    {{"name": "القيمة 1", "description": "شرح القيمة"}},
    ...
  ],
  "strategic_goals": [
    {{
      "goal": "الهدف الاستراتيجي",
      "description": "وصف تفصيلي",
      "focus_area": "المجال (مثل: مالي، عملاء، عمليات، تعلم)",
      "timeframe": "الإطار الزمني"
    }},
    ...
  ]
}}"""
        
        # Call AI using HuggingFace (FREE!)
        ai = AIManager.for_use_case('vision_mission')
        system_prompt = "أنت مستشار استراتيجي خبير متخصص في بناء الأطر الاستراتيجية للمؤسسات (الرؤية، الرسالة، القيم، الأهداف). قدّم إطاراً استراتيجياً ملهماً وقابلاً للتطبيق بصيغة JSON."
        response = ai.chat(
            prompt=prompt,
            system_prompt=system_prompt,
            response_format='json'
        )
        
        # Parse response
        framework_data = json.loads(response)
        
        # Save to database
        plan.vision_statement = framework_data.get('vision')
        plan.mission_statement = framework_data.get('mission')
        plan.core_values = json.dumps(framework_data.get('values', []), ensure_ascii=False)
        plan.strategic_goals = json.dumps(framework_data.get('strategic_goals', []), ensure_ascii=False)
        db.session.commit()
        
        return jsonify({'success': True, 'data': framework_data})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== KPIs GENERATION ====================

@strategic_planning_bp.route('/plan/<int:plan_id>/kpis')
@login_required
def manage_kpis(plan_id):
    """Manage KPIs for the strategic plan"""
    db = get_db()
    lang = get_lang()
    user_id = int(get_jwt_identity())
    
    # Authorization: Verify ownership
    plan = db.session.query(StrategicPlan).filter_by(id=plan_id).first_or_404()
    if plan.user_id != user_id:
        flash('غير مصرح لك بالوصول لهذه الخطة / Unauthorized access' if lang == 'ar' else 'Unauthorized access', 'danger')
        return redirect(url_for('strategic_planning_ai.index')), 403
    
    kpis = db.session.query(StrategicKPI).filter_by(plan_id=plan_id).all()
    
    return render_template('strategic_planning/kpis.html',
                         plan=plan,
                         kpis=kpis,
                         lang=lang)

@strategic_planning_bp.route('/plan/<int:plan_id>/generate-kpis', methods=['POST'])
@login_required
def generate_kpis(plan_id):
    """AI generates SMART KPIs for each strategic goal"""
    db = get_db()
    user_id = int(get_jwt_identity())
    
    # Authorization: Verify ownership
    plan = db.session.query(StrategicPlan).filter_by(id=plan_id).first_or_404()
    if plan.user_id != user_id:
        return jsonify({'success': False, 'error': 'Unauthorized access'}), 403
    
    try:
        strategic_goals = json.loads(plan.strategic_goals) if plan.strategic_goals else []
        
        prompt = f"""قم بتوليد مؤشرات أداء رئيسية (KPIs) قابلة للقياس لكل هدف استراتيجي:

**المؤسسة:** {plan.title}
**القطاع:** {plan.industry_sector}

**الأهداف الاستراتيجية:**
{chr(10).join(f'{i+1}. {g.get("goal")} ({g.get("focus_area")})' for i, g in enumerate(strategic_goals))}

لكل هدف، قم بتوليد 2-3 مؤشرات أداء SMART تشمل:
- اسم المؤشر
- وصف موجز
- الفئة (مالي، عملاء، عمليات داخلية، تعلم ونمو)
- وحدة القياس (%, عدد، مبلغ مالي، الخ)
- القيمة الأساسية (الحالية)
- القيمة المستهدفة
- دورية القياس (شهري، ربع سنوي، سنوي)
- الجهة المسؤولة

الرد بصيغة JSON:
{{
  "kpis": [
    {{
      "name": "اسم المؤشر",
      "description": "الوصف",
      "category": "الفئة",
      "measurement_unit": "الوحدة",
      "baseline_value": 0,
      "target_value": 100,
      "measurement_frequency": "شهري/ربع سنوي/سنوي",
      "responsible_party": "الإدارة المسؤولة",
      "data_source": "مصدر البيانات"
    }},
    ...
  ]
}}"""
        
        # Call AI using HuggingFace (FREE!)
        ai = AIManager.for_use_case('kpi_generation')
        system_prompt = "أنت مستشار استراتيجي خبير متخصص في تطوير مؤشرات الأداء الرئيسية (KPIs) بمعايير SMART. قدّم مؤشرات أداء قابلة للقياس والتطبيق بصيغة JSON."
        response = ai.chat(
            prompt=prompt,
            system_prompt=system_prompt,
            response_format='json'
        )
        
        # Parse response
        kpis_data = json.loads(response)
        
        # Create KPI records
        for kpi_data in kpis_data.get('kpis', []):
            kpi = StrategicKPI(
                plan_id=plan.id,
                name=kpi_data.get('name'),
                description=kpi_data.get('description'),
                category=kpi_data.get('category'),
                measurement_unit=kpi_data.get('measurement_unit'),
                baseline_value=kpi_data.get('baseline_value', 0),
                target_value=kpi_data.get('target_value', 0),
                current_value=kpi_data.get('baseline_value', 0),
                measurement_frequency=kpi_data.get('measurement_frequency'),
                responsible_party=kpi_data.get('responsible_party'),
                data_source=kpi_data.get('data_source', ''),
                status='active'
            )
            db.session.add(kpi)
        db.session.commit()
        
        return jsonify({'success': True, 'kpis_count': len(kpis_data.get('kpis', []))})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== PLAN DASHBOARD ====================

@strategic_planning_bp.route('/plan/<int:plan_id>/dashboard')
@login_required
def plan_dashboard(plan_id):
    """Complete strategic plan dashboard with all components"""
    db = get_db()
    lang = get_lang()
    user_id = int(get_jwt_identity())
    
    # Authorization: Verify ownership
    plan = db.session.query(StrategicPlan).filter_by(id=plan_id).first_or_404()
    if plan.user_id != user_id:
        flash('غير مصرح لك بالوصول لهذه الخطة / Unauthorized access' if lang == 'ar' else 'Unauthorized access', 'danger')
        return redirect(url_for('strategic_planning_ai.index')), 403
    
    kpis = db.session.query(StrategicKPI).filter_by(plan_id=plan_id).all()
    initiatives = db.session.query(StrategicInitiative).filter_by(plan_id=plan_id).all()
    
    # Parse JSON fields
    swot = json.loads(plan.swot_analysis) if plan.swot_analysis else {}
    pestel = json.loads(plan.pestel_analysis) if plan.pestel_analysis else {}
    values = json.loads(plan.core_values) if plan.core_values else []
    goals = json.loads(plan.strategic_goals) if plan.strategic_goals else []
    
    return render_template('strategic_planning/dashboard.html',
                         plan=plan,
                         swot=swot,
                         pestel=pestel,
                         values=values,
                         goals=goals,
                         kpis=kpis,
                         initiatives=initiatives,
                         lang=lang)

# ==================== EXPORT & REPORTS ====================

@strategic_planning_bp.route('/plan/<int:plan_id>/export-pdf')
@login_required
def export_pdf(plan_id):
    """Export strategic plan as PDF report with Arabic support"""
    from flask import make_response
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT, TA_JUSTIFY
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from datetime import datetime
    import arabic_reshaper
    from bidi.algorithm import get_display
    
    db = get_db()
    lang = get_lang()
    user_id = int(get_jwt_identity())
    
    # Authorization: Verify ownership
    plan = db.session.query(StrategicPlan).filter_by(id=plan_id).first_or_404()
    if plan.user_id != user_id:
        flash('غير مصرح لك بالوصول لهذه الخطة / Unauthorized access', 'danger')
        return redirect(url_for('strategic_planning_ai.index')), 403
    
    def prep_text(text):
        """Prepare text for PDF - Arabic reshaping if needed"""
        if not text:
            return ""
        try:
            # Check if text contains Arabic characters
            if any('\u0600' <= c <= '\u06FF' for c in text):
                reshaped = arabic_reshaper.reshape(text)
                return get_display(reshaped)
            return text
        except:
            return text
    
    try:
        # Create PDF in memory
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
        
        # Build content
        story = []
        styles = getSampleStyleSheet()
        
        # Define heading style for sections
        heading_style = ParagraphStyle('Heading', parent=styles['Heading2'], fontSize=14, textColor=colors.HexColor('#1e3a8a'), spaceAfter=6, alignment=TA_RIGHT if lang == 'ar' else TA_LEFT)
        normal_style = ParagraphStyle('NormalText', parent=styles['Normal'], alignment=TA_RIGHT if lang == 'ar' else TA_LEFT)
        
        # Title
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#1e3a8a'), spaceAfter=12, alignment=TA_CENTER)
        story.append(Paragraph(prep_text(plan.title), title_style))
        story.append(Spacer(1, 0.3*cm))
        
        # Planning Period
        period_style = ParagraphStyle('Period', parent=styles['Normal'], fontSize=12, textColor=colors.grey, alignment=TA_CENTER)
        period_label = prep_text("فترة التخطيط:" if lang == 'ar' else "Planning Period:")
        story.append(Paragraph(f"{period_label} {plan.planning_period}", period_style))
        story.append(Spacer(1, 0.5*cm))
        
        # Vision & Mission
        if plan.vision_statement or plan.mission_statement:
            
            if plan.vision_statement:
                vision_label = prep_text("الرؤية" if lang == 'ar' else "Vision")
                story.append(Paragraph(f"<b>{vision_label}</b>", heading_style))
                story.append(Paragraph(prep_text(plan.vision_statement), normal_style))
                story.append(Spacer(1, 0.3*cm))
            
            if plan.mission_statement:
                mission_label = prep_text("الرسالة" if lang == 'ar' else "Mission")
                story.append(Paragraph(f"<b>{mission_label}</b>", heading_style))
                story.append(Paragraph(prep_text(plan.mission_statement), normal_style))
                story.append(Spacer(1, 0.5*cm))
        
        # Strategic Goals
        goals = json.loads(plan.strategic_goals) if plan.strategic_goals else []
        if goals:
            goals_label = prep_text("الأهداف الاستراتيجية" if lang == 'ar' else "Strategic Goals")
            story.append(Paragraph(f"<b>{goals_label}</b>", heading_style))
            for i, goal in enumerate(goals, 1):
                goal_text = prep_text(goal.get('goal', ''))
                goal_desc = prep_text(goal.get('description', ''))
                story.append(Paragraph(f"<b>{i}. {goal_text}</b>", normal_style))
                story.append(Paragraph(goal_desc, normal_style))
                story.append(Spacer(1, 0.2*cm))
            story.append(Spacer(1, 0.3*cm))
        
        # SWOT Analysis
        swot = json.loads(plan.swot_analysis) if plan.swot_analysis else {}
        if swot:
            story.append(PageBreak())
            swot_label = prep_text("تحليل SWOT" if lang == 'ar' else "SWOT Analysis")
            story.append(Paragraph(f"<b>{swot_label}</b>", heading_style))
            
            strengths_label = prep_text("نقاط القوة" if lang == 'ar' else "Strengths")
            weaknesses_label = prep_text("نقاط الضعف" if lang == 'ar' else "Weaknesses")
            opportunities_label = prep_text("الفرص" if lang == 'ar' else "Opportunities")
            threats_label = prep_text("التهديدات" if lang == 'ar' else "Threats")
            
            swot_data = [
                [Paragraph(f"<b>{strengths_label}</b>", normal_style), Paragraph(f"<b>{weaknesses_label}</b>", normal_style)],
                [Paragraph("<br/>".join(f"• {prep_text(s)}" for s in swot.get('strengths', [])), normal_style), 
                 Paragraph("<br/>".join(f"• {prep_text(w)}" for w in swot.get('weaknesses', [])), normal_style)],
                [Paragraph(f"<b>{opportunities_label}</b>", normal_style), Paragraph(f"<b>{threats_label}</b>", normal_style)],
                [Paragraph("<br/>".join(f"• {prep_text(o)}" for o in swot.get('opportunities', [])), normal_style), 
                 Paragraph("<br/>".join(f"• {prep_text(t)}" for t in swot.get('threats', [])), normal_style)]
            ]
            
            swot_table = Table(swot_data, colWidths=[8*cm, 8*cm])
            swot_table.setStyle(TableStyle([
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e0e7ff')),
                ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#e0e7ff')),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('PADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(swot_table)
            story.append(Spacer(1, 0.5*cm))
        
        # PESTEL Analysis
        pestel = json.loads(plan.pestel_analysis) if plan.pestel_analysis else {}
        if pestel:
            story.append(PageBreak())
            pestel_label = prep_text("تحليل PESTEL" if lang == 'ar' else "PESTEL Analysis")
            story.append(Paragraph(f"<b>{pestel_label}</b>", heading_style))
            
            categories = [
                ("سياسية" if lang == 'ar' else "Political", pestel.get('political', [])),
                ("اقتصادية" if lang == 'ar' else "Economic", pestel.get('economic', [])),
                ("اجتماعية" if lang == 'ar' else "Social", pestel.get('social', [])),
                ("تقنية" if lang == 'ar' else "Technological", pestel.get('technological', [])),
                ("بيئية" if lang == 'ar' else "Environmental", pestel.get('environmental', [])),
                ("قانونية" if lang == 'ar' else "Legal", pestel.get('legal', []))
            ]
            
            for category, items in categories:
                if items:
                    cat_label = prep_text(category)
                    story.append(Paragraph(f"<b>{cat_label}:</b>", normal_style))
                    items_text = "<br/>".join(f"• {prep_text(item)}" for item in items)
                    story.append(Paragraph(items_text, normal_style))
                    story.append(Spacer(1, 0.2*cm))
        
        # KPIs
        kpis = db.session.query(StrategicKPI).filter_by(plan_id=plan_id).all()
        if kpis:
            story.append(PageBreak())
            kpi_label = prep_text("مؤشرات الأداء الرئيسية" if lang == 'ar' else "Key Performance Indicators (KPIs)")
            story.append(Paragraph(f"<b>{kpi_label}</b>", heading_style))
            
            # Simple list instead of table for better Arabic support
            for kpi in kpis:
                kpi_name = prep_text(kpi.name)
                kpi_cat = prep_text(kpi.category or '')
                story.append(Paragraph(f"<b>• {kpi_name}</b> ({kpi_cat})", normal_style))
                current_label = prep_text("الحالي:" if lang == 'ar' else "Current:")
                target_label = prep_text("المستهدف:" if lang == 'ar' else "Target:")
                story.append(Paragraph(f"  {current_label} {kpi.current_value} {kpi.measurement_unit} | {target_label} {kpi.target_value} {kpi.measurement_unit}", normal_style))
                story.append(Spacer(1, 0.3*cm))
        
        # Footer
        story.append(Spacer(1, 1*cm))
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
        footer_text = prep_text(f"تم الإنشاء في {datetime.now().strftime('%Y-%m-%d %H:%M')} بواسطة منصة Mcidia" if lang == 'ar' else f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')} by Mcidia Platform")
        story.append(Paragraph(footer_text, footer_style))
        
        # Build PDF
        doc.build(story)
        
        # Prepare response
        buffer.seek(0)
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename="strategic_plan_{plan_id}.pdf"'
        
        return response
        
    except Exception as e:
        current_app.logger.error(f"PDF generation error: {str(e)}")
        flash(f'خطأ في إنشاء PDF / Error generating PDF: {str(e)}', 'danger')
        return redirect(url_for('strategic_planning_ai.plan_dashboard', plan_id=plan_id))

@strategic_planning_bp.route('/plan/<int:plan_id>/export-excel')
@login_required
def export_excel(plan_id):
    """Export strategic plan as comprehensive Excel report with Arabic support"""
    db = get_db()
    lang = get_lang()
    user_id = int(get_jwt_identity())
    
    # Authorization check
    plan = db.session.query(StrategicPlan).filter_by(id=plan_id).first_or_404()
    if plan.user_id != user_id:
        flash('غير مصرح لك بالوصول لهذه الخطة / Unauthorized access', 'danger')
        return redirect(url_for('strategic_planning_ai.index')), 403
    
    try:
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Define styles
        header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
        header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        subheader_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
        subheader_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        section_fill = PatternFill(start_color="e0e7ff", end_color="e0e7ff", fill_type="solid")
        section_font = Font(name='Arial', size=10, bold=True)
        normal_font = Font(name='Arial', size=10)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # RTL alignment for Arabic
        rtl_alignment = Alignment(horizontal='right', vertical='top', wrap_text=True, readingOrder=2)
        ltr_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # ===== Sheet 1: Overview (نظرة عامة) =====
        ws_overview = wb.create_sheet(title="نظرة عامة - Overview" if lang == 'ar' else "Overview")
        ws_overview.sheet_view.rightToLeft = (lang == 'ar')
        
        # Title
        ws_overview['A1'] = plan.title
        ws_overview['A1'].font = Font(name='Arial', size=16, bold=True)
        ws_overview['A1'].alignment = rtl_alignment if lang == 'ar' else ltr_alignment
        ws_overview.merge_cells('A1:D1')
        
        # Planning Period
        row = 3
        ws_overview[f'A{row}'] = "فترة التخطيط:" if lang == 'ar' else "Planning Period:"
        ws_overview[f'A{row}'].font = section_font
        ws_overview[f'B{row}'] = plan.planning_period
        ws_overview[f'B{row}'].font = normal_font
        
        # Status
        row += 1
        ws_overview[f'A{row}'] = "الحالة:" if lang == 'ar' else "Status:"
        ws_overview[f'A{row}'].font = section_font
        ws_overview[f'B{row}'] = plan.status
        ws_overview[f'B{row}'].font = normal_font
        
        # Organization Info
        row += 2
        ws_overview[f'A{row}'] = "معلومات المنظمة - Organization Information" if lang == 'ar' else "Organization Information"
        ws_overview[f'A{row}'].font = subheader_font
        ws_overview[f'A{row}'].fill = subheader_fill
        ws_overview[f'A{row}'].alignment = center_alignment
        ws_overview.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws_overview[f'A{row}'] = "اسم المنظمة:" if lang == 'ar' else "Organization Name:"
        ws_overview[f'A{row}'].font = section_font
        ws_overview[f'B{row}'] = plan.organization_name or ""
        ws_overview[f'B{row}'].font = normal_font
        ws_overview[f'B{row}'].alignment = rtl_alignment if lang == 'ar' else ltr_alignment
        
        row += 1
        ws_overview[f'A{row}'] = "القطاع:" if lang == 'ar' else "Industry Sector:"
        ws_overview[f'A{row}'].font = section_font
        ws_overview[f'B{row}'] = plan.industry_sector or ""
        ws_overview[f'B{row}'].font = normal_font
        ws_overview[f'B{row}'].alignment = rtl_alignment if lang == 'ar' else ltr_alignment
        
        row += 1
        ws_overview[f'A{row}'] = "عدد الموظفين:" if lang == 'ar' else "Employee Count:"
        ws_overview[f'A{row}'].font = section_font
        ws_overview[f'B{row}'] = plan.employee_count or ""
        ws_overview[f'B{row}'].font = normal_font
        
        # Vision & Mission
        row += 2
        if plan.vision_statement:
            ws_overview[f'A{row}'] = "الرؤية - Vision" if lang == 'ar' else "Vision"
            ws_overview[f'A{row}'].font = subheader_font
            ws_overview[f'A{row}'].fill = subheader_fill
            ws_overview[f'A{row}'].alignment = center_alignment
            ws_overview.merge_cells(f'A{row}:D{row}')
            
            row += 1
            ws_overview[f'A{row}'] = plan.vision_statement
            ws_overview[f'A{row}'].font = normal_font
            ws_overview[f'A{row}'].alignment = rtl_alignment if lang == 'ar' else ltr_alignment
            ws_overview.merge_cells(f'A{row}:D{row}')
            ws_overview.row_dimensions[row].height = 50
            row += 2
        
        if plan.mission_statement:
            ws_overview[f'A{row}'] = "الرسالة - Mission" if lang == 'ar' else "Mission"
            ws_overview[f'A{row}'].font = subheader_font
            ws_overview[f'A{row}'].fill = subheader_fill
            ws_overview[f'A{row}'].alignment = center_alignment
            ws_overview.merge_cells(f'A{row}:D{row}')
            
            row += 1
            ws_overview[f'A{row}'] = plan.mission_statement
            ws_overview[f'A{row}'].font = normal_font
            ws_overview[f'A{row}'].alignment = rtl_alignment if lang == 'ar' else ltr_alignment
            ws_overview.merge_cells(f'A{row}:D{row}')
            ws_overview.row_dimensions[row].height = 50
        
        # Set column widths
        ws_overview.column_dimensions['A'].width = 25
        ws_overview.column_dimensions['B'].width = 40
        ws_overview.column_dimensions['C'].width = 40
        ws_overview.column_dimensions['D'].width = 40
        
        # ===== Sheet 2: Core Values (القيم الجوهرية) =====
        core_values = json.loads(plan.core_values) if plan.core_values else []
        if core_values:
            ws_values = wb.create_sheet(title="القيم الجوهرية - Core Values" if lang == 'ar' else "Core Values")
            ws_values.sheet_view.rightToLeft = (lang == 'ar')
            
            # Header
            ws_values['A1'] = "القيم الجوهرية - Core Values" if lang == 'ar' else "Core Values"
            ws_values['A1'].font = header_font
            ws_values['A1'].fill = header_fill
            ws_values['A1'].alignment = center_alignment
            ws_values.merge_cells('A1:C1')
            
            # Column headers
            ws_values['A3'] = "#"
            ws_values['B3'] = "القيمة - Value" if lang == 'ar' else "Value"
            ws_values['C3'] = "الوصف - Description" if lang == 'ar' else "Description"
            
            for col in ['A', 'B', 'C']:
                ws_values[f'{col}3'].font = section_font
                ws_values[f'{col}3'].fill = section_fill
                ws_values[f'{col}3'].alignment = center_alignment
                ws_values[f'{col}3'].border = border
            
            # Data
            row = 4
            for idx, value in enumerate(core_values, 1):
                ws_values[f'A{row}'] = idx
                ws_values[f'B{row}'] = value.get('value', '')
                ws_values[f'C{row}'] = value.get('description', '')
                
                for col in ['A', 'B', 'C']:
                    ws_values[f'{col}{row}'].font = normal_font
                    ws_values[f'{col}{row}'].alignment = rtl_alignment if lang == 'ar' and col != 'A' else center_alignment if col == 'A' else ltr_alignment
                    ws_values[f'{col}{row}'].border = border
                
                ws_values.row_dimensions[row].height = 40
                row += 1
            
            ws_values.column_dimensions['A'].width = 8
            ws_values.column_dimensions['B'].width = 30
            ws_values.column_dimensions['C'].width = 60
        
        # ===== Sheet 3: Strategic Goals (الأهداف الاستراتيجية) =====
        goals = json.loads(plan.strategic_goals) if plan.strategic_goals else []
        if goals:
            ws_goals = wb.create_sheet(title="الأهداف - Strategic Goals" if lang == 'ar' else "Strategic Goals")
            ws_goals.sheet_view.rightToLeft = (lang == 'ar')
            
            # Header
            ws_goals['A1'] = "الأهداف الاستراتيجية - Strategic Goals" if lang == 'ar' else "Strategic Goals"
            ws_goals['A1'].font = header_font
            ws_goals['A1'].fill = header_fill
            ws_goals['A1'].alignment = center_alignment
            ws_goals.merge_cells('A1:D1')
            
            # Column headers
            ws_goals['A3'] = "#"
            ws_goals['B3'] = "الهدف - Goal" if lang == 'ar' else "Goal"
            ws_goals['C3'] = "الوصف - Description" if lang == 'ar' else "Description"
            ws_goals['D3'] = "الإطار الزمني - Timeline" if lang == 'ar' else "Timeline"
            
            for col in ['A', 'B', 'C', 'D']:
                ws_goals[f'{col}3'].font = section_font
                ws_goals[f'{col}3'].fill = section_fill
                ws_goals[f'{col}3'].alignment = center_alignment
                ws_goals[f'{col}3'].border = border
            
            # Data
            row = 4
            for idx, goal in enumerate(goals, 1):
                ws_goals[f'A{row}'] = idx
                ws_goals[f'B{row}'] = goal.get('goal', '')
                ws_goals[f'C{row}'] = goal.get('description', '')
                ws_goals[f'D{row}'] = goal.get('timeline', '')
                
                for col in ['A', 'B', 'C', 'D']:
                    ws_goals[f'{col}{row}'].font = normal_font
                    ws_goals[f'{col}{row}'].alignment = rtl_alignment if lang == 'ar' and col != 'A' else center_alignment if col == 'A' else ltr_alignment
                    ws_goals[f'{col}{row}'].border = border
                
                ws_goals.row_dimensions[row].height = 40
                row += 1
            
            ws_goals.column_dimensions['A'].width = 8
            ws_goals.column_dimensions['B'].width = 35
            ws_goals.column_dimensions['C'].width = 50
            ws_goals.column_dimensions['D'].width = 20
        
        # ===== Sheet 4: SWOT Analysis =====
        swot = json.loads(plan.swot_analysis) if plan.swot_analysis else {}
        if swot:
            ws_swot = wb.create_sheet(title="تحليل SWOT" if lang == 'ar' else "SWOT Analysis")
            ws_swot.sheet_view.rightToLeft = (lang == 'ar')
            
            # Header
            ws_swot['A1'] = "تحليل SWOT" if lang == 'ar' else "SWOT Analysis"
            ws_swot['A1'].font = header_font
            ws_swot['A1'].fill = header_fill
            ws_swot['A1'].alignment = center_alignment
            ws_swot.merge_cells('A1:B1')
            
            # Strengths
            row = 3
            ws_swot[f'A{row}'] = "نقاط القوة - Strengths" if lang == 'ar' else "Strengths"
            ws_swot[f'A{row}'].font = subheader_font
            ws_swot[f'A{row}'].fill = subheader_fill
            ws_swot[f'A{row}'].alignment = center_alignment
            ws_swot.merge_cells(f'A{row}:B{row}')
            
            for idx, item in enumerate(swot.get('strengths', []), 1):
                row += 1
                ws_swot[f'A{row}'] = idx
                ws_swot[f'B{row}'] = item
                ws_swot[f'A{row}'].alignment = center_alignment
                ws_swot[f'B{row}'].alignment = rtl_alignment if lang == 'ar' else ltr_alignment
                ws_swot[f'B{row}'].font = normal_font
            
            # Weaknesses
            row += 2
            ws_swot[f'A{row}'] = "نقاط الضعف - Weaknesses" if lang == 'ar' else "Weaknesses"
            ws_swot[f'A{row}'].font = subheader_font
            ws_swot[f'A{row}'].fill = subheader_fill
            ws_swot[f'A{row}'].alignment = center_alignment
            ws_swot.merge_cells(f'A{row}:B{row}')
            
            for idx, item in enumerate(swot.get('weaknesses', []), 1):
                row += 1
                ws_swot[f'A{row}'] = idx
                ws_swot[f'B{row}'] = item
                ws_swot[f'A{row}'].alignment = center_alignment
                ws_swot[f'B{row}'].alignment = rtl_alignment if lang == 'ar' else ltr_alignment
                ws_swot[f'B{row}'].font = normal_font
            
            # Opportunities
            row += 2
            ws_swot[f'A{row}'] = "الفرص - Opportunities" if lang == 'ar' else "Opportunities"
            ws_swot[f'A{row}'].font = subheader_font
            ws_swot[f'A{row}'].fill = subheader_fill
            ws_swot[f'A{row}'].alignment = center_alignment
            ws_swot.merge_cells(f'A{row}:B{row}')
            
            for idx, item in enumerate(swot.get('opportunities', []), 1):
                row += 1
                ws_swot[f'A{row}'] = idx
                ws_swot[f'B{row}'] = item
                ws_swot[f'A{row}'].alignment = center_alignment
                ws_swot[f'B{row}'].alignment = rtl_alignment if lang == 'ar' else ltr_alignment
                ws_swot[f'B{row}'].font = normal_font
            
            # Threats
            row += 2
            ws_swot[f'A{row}'] = "التهديدات - Threats" if lang == 'ar' else "Threats"
            ws_swot[f'A{row}'].font = subheader_font
            ws_swot[f'A{row}'].fill = subheader_fill
            ws_swot[f'A{row}'].alignment = center_alignment
            ws_swot.merge_cells(f'A{row}:B{row}')
            
            for idx, item in enumerate(swot.get('threats', []), 1):
                row += 1
                ws_swot[f'A{row}'] = idx
                ws_swot[f'B{row}'] = item
                ws_swot[f'A{row}'].alignment = center_alignment
                ws_swot[f'B{row}'].alignment = rtl_alignment if lang == 'ar' else ltr_alignment
                ws_swot[f'B{row}'].font = normal_font
            
            ws_swot.column_dimensions['A'].width = 8
            ws_swot.column_dimensions['B'].width = 80
        
        # ===== Sheet 5: PESTEL Analysis =====
        pestel = json.loads(plan.pestel_analysis) if plan.pestel_analysis else {}
        if pestel:
            ws_pestel = wb.create_sheet(title="تحليل PESTEL" if lang == 'ar' else "PESTEL Analysis")
            ws_pestel.sheet_view.rightToLeft = (lang == 'ar')
            
            # Header
            ws_pestel['A1'] = "تحليل PESTEL" if lang == 'ar' else "PESTEL Analysis"
            ws_pestel['A1'].font = header_font
            ws_pestel['A1'].fill = header_fill
            ws_pestel['A1'].alignment = center_alignment
            ws_pestel.merge_cells('A1:B1')
            
            categories = [
                ("سياسية - Political" if lang == 'ar' else "Political", pestel.get('political', [])),
                ("اقتصادية - Economic" if lang == 'ar' else "Economic", pestel.get('economic', [])),
                ("اجتماعية - Social" if lang == 'ar' else "Social", pestel.get('social', [])),
                ("تقنية - Technological" if lang == 'ar' else "Technological", pestel.get('technological', [])),
                ("بيئية - Environmental" if lang == 'ar' else "Environmental", pestel.get('environmental', [])),
                ("قانونية - Legal" if lang == 'ar' else "Legal", pestel.get('legal', []))
            ]
            
            row = 3
            for category, items in categories:
                ws_pestel[f'A{row}'] = category
                ws_pestel[f'A{row}'].font = subheader_font
                ws_pestel[f'A{row}'].fill = subheader_fill
                ws_pestel[f'A{row}'].alignment = center_alignment
                ws_pestel.merge_cells(f'A{row}:B{row}')
                
                for idx, item in enumerate(items, 1):
                    row += 1
                    ws_pestel[f'A{row}'] = idx
                    ws_pestel[f'B{row}'] = item
                    ws_pestel[f'A{row}'].alignment = center_alignment
                    ws_pestel[f'B{row}'].alignment = rtl_alignment if lang == 'ar' else ltr_alignment
                    ws_pestel[f'B{row}'].font = normal_font
                
                row += 2
            
            ws_pestel.column_dimensions['A'].width = 8
            ws_pestel.column_dimensions['B'].width = 80
        
        # ===== Sheet 6: KPIs =====
        kpis = db.session.query(StrategicKPI).filter_by(plan_id=plan_id).all()
        if kpis:
            ws_kpis = wb.create_sheet(title="مؤشرات الأداء - KPIs" if lang == 'ar' else "KPIs")
            ws_kpis.sheet_view.rightToLeft = (lang == 'ar')
            
            # Header
            ws_kpis['A1'] = "مؤشرات الأداء الرئيسية - KPIs" if lang == 'ar' else "Key Performance Indicators"
            ws_kpis['A1'].font = header_font
            ws_kpis['A1'].fill = header_fill
            ws_kpis['A1'].alignment = center_alignment
            ws_kpis.merge_cells('A1:H1')
            
            # Column headers
            headers = [
                "#",
                "اسم المؤشر - KPI Name" if lang == 'ar' else "KPI Name",
                "الفئة - Category" if lang == 'ar' else "Category",
                "القيمة الحالية - Current" if lang == 'ar' else "Current Value",
                "القيمة المستهدفة - Target" if lang == 'ar' else "Target Value",
                "وحدة القياس - Unit" if lang == 'ar' else "Unit",
                "الحالة - Status" if lang == 'ar' else "Status",
                "التقدم % - Progress %" if lang == 'ar' else "Progress %"
            ]
            
            for idx, header in enumerate(headers, 1):
                col_letter = get_column_letter(idx)
                ws_kpis[f'{col_letter}3'] = header
                ws_kpis[f'{col_letter}3'].font = section_font
                ws_kpis[f'{col_letter}3'].fill = section_fill
                ws_kpis[f'{col_letter}3'].alignment = center_alignment
                ws_kpis[f'{col_letter}3'].border = border
            
            # Data
            row = 4
            for idx, kpi in enumerate(kpis, 1):
                progress = 0
                if kpi.target_value and kpi.target_value != 0:
                    progress = round((kpi.current_value / kpi.target_value) * 100, 1)
                
                ws_kpis[f'A{row}'] = idx
                ws_kpis[f'B{row}'] = kpi.name
                ws_kpis[f'C{row}'] = kpi.category or ""
                ws_kpis[f'D{row}'] = kpi.current_value
                ws_kpis[f'E{row}'] = kpi.target_value
                ws_kpis[f'F{row}'] = kpi.measurement_unit
                ws_kpis[f'G{row}'] = kpi.status
                ws_kpis[f'H{row}'] = f"{progress}%"
                
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                    ws_kpis[f'{col}{row}'].font = normal_font
                    ws_kpis[f'{col}{row}'].alignment = rtl_alignment if lang == 'ar' and col in ['B', 'C', 'F', 'G'] else center_alignment
                    ws_kpis[f'{col}{row}'].border = border
                
                # Color code progress
                if progress >= 80:
                    ws_kpis[f'H{row}'].fill = PatternFill(start_color="22c55e", end_color="22c55e", fill_type="solid")
                elif progress >= 50:
                    ws_kpis[f'H{row}'].fill = PatternFill(start_color="eab308", end_color="eab308", fill_type="solid")
                else:
                    ws_kpis[f'H{row}'].fill = PatternFill(start_color="ef4444", end_color="ef4444", fill_type="solid")
                
                row += 1
            
            ws_kpis.column_dimensions['A'].width = 6
            ws_kpis.column_dimensions['B'].width = 35
            ws_kpis.column_dimensions['C'].width = 20
            ws_kpis.column_dimensions['D'].width = 15
            ws_kpis.column_dimensions['E'].width = 15
            ws_kpis.column_dimensions['F'].width = 15
            ws_kpis.column_dimensions['G'].width = 15
            ws_kpis.column_dimensions['H'].width = 12
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Prepare response
        response = make_response(excel_buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="strategic_plan_{plan_id}.xlsx"'
        
        return response
        
    except Exception as e:
        current_app.logger.error(f"Excel generation error: {str(e)}")
        flash(f'خطأ في إنشاء ملف Excel / Error generating Excel: {str(e)}', 'danger')
        return redirect(url_for('strategic_planning_ai.plan_dashboard', plan_id=plan_id))

@strategic_planning_bp.route('/plan/<int:plan_id>/delete', methods=['POST'])
@login_required
def delete_plan(plan_id):
    """Delete a strategic plan and all related data"""
    db = get_db()
    lang = get_lang()
    user_id = int(get_jwt_identity())
    
    # Authorization: Verify ownership
    plan = db.session.query(StrategicPlan).filter_by(id=plan_id).first_or_404()
    if plan.user_id != user_id:
        flash('غير مصرح لك بحذف هذه الخطة / Unauthorized to delete this plan', 'danger')
        return redirect(url_for('strategic_planning_ai.index')), 403
    
    try:
        plan_title = plan.title
        
        # Delete the plan (cascade will delete related KPIs and initiatives)
        db.session.delete(plan)
        db.session.commit()
        
        flash(f'تم حذف الخطة "{plan_title}" بنجاح / Plan "{plan_title}" deleted successfully' if lang == 'ar' else f'Plan "{plan_title}" deleted successfully', 'success')
        return redirect(url_for('strategic_planning_ai.index'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'خطأ في حذف الخطة: {str(e)} / Error deleting plan: {str(e)}' if lang == 'ar' else f'Error deleting plan: {str(e)}', 'danger')
        return redirect(url_for('strategic_planning_ai.plan_dashboard', plan_id=plan_id))
