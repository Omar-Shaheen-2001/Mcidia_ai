from flask import Blueprint, render_template, request, flash, redirect, url_for, session, jsonify, send_file
from flask_jwt_extended import get_jwt_identity
from utils.decorators import login_required, role_required
from models import User, Role, SubscriptionPlan, Project, Transaction, AILog, Organization, OrganizationMembership
from flask import current_app
from werkzeug.security import generate_password_hash
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
try:
    from weasyprint import HTML, CSS
    WEASYPRINT_AVAILABLE = True
except (ImportError, OSError):
    WEASYPRINT_AVAILABLE = False
    HTML, CSS = None, None

users_bp = Blueprint('users', __name__, url_prefix='/users')

def get_db():
    return current_app.extensions['sqlalchemy']

def get_lang():
    return session.get('language', 'ar')

@users_bp.route('/export/excel', methods=['GET'])
@login_required
@role_required('system_admin')
def export_excel():
    """Export users data to Excel"""
    db = get_db()
    lang = get_lang()
    
    # Get users with filters applied
    query = db.session.query(User)
    users = query.order_by(User.created_at.desc()).all()
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'المستخدمون' if lang == 'ar' else 'Users'
    
    # Define styles
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Headers
    headers = ['ID', 'اسم المستخدم' if lang == 'ar' else 'Username', 
               'البريد الإلكتروني' if lang == 'ar' else 'Email',
               'رقم الهاتف' if lang == 'ar' else 'Phone',
               'الشركة' if lang == 'ar' else 'Company',
               'الدور' if lang == 'ar' else 'Role',
               'الخطة' if lang == 'ar' else 'Plan',
               'الحالة' if lang == 'ar' else 'Status',
               'آخر دخول' if lang == 'ar' else 'Last Login',
               'عنوان IP' if lang == 'ar' else 'Last IP',
               'تاريخ التسجيل' if lang == 'ar' else 'Join Date']
    
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # Add data rows
    for user in users:
        row = [
            user.id,
            user.username,
            user.email,
            user.phone or '-',
            user.company_name or '-',
            user.role or '-',
            user.subscription_plan or '-',
            'نشط' if lang == 'ar' and user.is_active else 'Active' if user.is_active else 'غير نشط' if lang == 'ar' else 'Inactive',
            user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else '-',
            user.last_login_ip or '-',
            user.created_at.strftime('%Y-%m-%d') if user.created_at else '-'
        ]
        ws.append(row)
        
        # Style data cells
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if cell.column <= 1 else 'left', vertical='center', wrap_text=True)
    
    # Adjust column widths
    column_widths = [8, 15, 20, 12, 15, 12, 12, 10, 16, 14, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'users_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


@users_bp.route('/export/pdf', methods=['GET'])
@login_required
@role_required('system_admin')
def export_pdf():
    """Export users data to PDF with Arabic support"""
    try:
        db = get_db()
        lang = get_lang()
        
        # Get users
        query = db.session.query(User)
        users = query.order_by(User.created_at.desc()).all()
        
        # Create HTML table for PDF
        timestamp = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
        title = 'تقرير المستخدمين' if lang == 'ar' else 'Users Report'
        
        # Table headers
        headers = {
            'ar': ['معرّف', 'اسم المستخدم', 'البريد الإلكتروني', 'الدور', 'الخطة', 'الحالة', 'تاريخ التسجيل'],
            'en': ['ID', 'Username', 'Email', 'Role', 'Plan', 'Status', 'Join Date']
        }
        
        status_text = {
            'ar': {'active': 'نشط', 'inactive': 'غير نشط'},
            'en': {'active': 'Active', 'inactive': 'Inactive'}
        }
        
        # Build table rows
        rows_html = ''
        for user in users[:100]:
            status = status_text[lang]['active'] if user.is_active else status_text[lang]['inactive']
            status_class = 'status-active' if user.is_active else 'status-inactive'
            rows_html += f"""
            <tr>
                <td>{user.id}</td>
                <td>{user.username[:15]}</td>
                <td>{user.email[:18]}</td>
                <td>{user.role or '-'}</td>
                <td>{user.subscription_plan or '-'}</td>
                <td><span class="{status_class}">{status}</span></td>
                <td>{user.created_at.strftime('%Y-%m-%d') if user.created_at else '-'}</td>
            </tr>
            """
        
        # Build HTML
        html_content = f"""
        <!DOCTYPE html>
        <html dir="{'rtl' if lang == 'ar' else 'ltr'}">
        <head>
            <meta charset="UTF-8">
            <style>
                @page {{
                    size: A4;
                    margin: 0.5cm;
                }}
                * {{
                    margin: 0;
                    padding: 0;
                    box-sizing: border-box;
                }}
                body {{
                    font-family: 'DejaVu Sans', sans-serif;
                    margin: 0;
                    padding: 8px;
                    direction: {'rtl' if lang == 'ar' else 'ltr'};
                    background-color: #fff;
                    font-size: 9pt;
                    line-height: 1.3;
                }}
                .header {{
                    text-align: center;
                    margin-bottom: 12px;
                    border-bottom: 2px solid #366092;
                    padding-bottom: 8px;
                }}
                h1 {{
                    color: #366092;
                    font-size: 14pt;
                    margin: 0;
                    font-weight: bold;
                }}
                .report-date {{
                    color: #666;
                    font-size: 8pt;
                    margin-top: 2px;
                }}
                .stats {{
                    display: flex;
                    justify-content: space-around;
                    margin-bottom: 10px;
                    background-color: #f5f5f5;
                    padding: 6px;
                    border: 1px solid #ddd;
                }}
                .stat-item {{
                    text-align: center;
                    flex: 1;
                }}
                .stat-label {{
                    color: #666;
                    font-size: 7pt;
                    display: block;
                    margin-bottom: 2px;
                }}
                .stat-value {{
                    color: #366092;
                    font-weight: bold;
                    font-size: 11pt;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin: 8px 0;
                    font-size: 8pt;
                }}
                th {{
                    background-color: #366092;
                    color: #fff;
                    padding: 5px 3px;
                    text-align: center;
                    font-weight: bold;
                    border: 1px solid #2c5aa0;
                    vertical-align: middle;
                    font-size: 8pt;
                }}
                td {{
                    padding: 4px 3px;
                    border: 1px solid #ddd;
                    text-align: center;
                    vertical-align: middle;
                    font-size: 8pt;
                    word-wrap: break-word;
                }}
                tbody tr:nth-child(odd) {{
                    background-color: #f9f9f9;
                }}
                tbody tr:nth-child(even) {{
                    background-color: #fff;
                }}
                .status-active {{
                    color: #1b5e20;
                    font-weight: bold;
                }}
                .status-inactive {{
                    color: #c62828;
                    font-weight: bold;
                }}
                .footer {{
                    text-align: center;
                    color: #999;
                    font-size: 7pt;
                    margin-top: 10px;
                    padding-top: 6px;
                    border-top: 1px solid #ddd;
                }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>{title}</h1>
                <div class="report-date">{timestamp}</div>
            </div>
            
            <div class="stats">
                <div class="stats-row">
                    <div class="stat-item">
                        <span class="stat-label">{'إجمالي المستخدمين' if lang == 'ar' else 'Total Users'}</span>
                        <span class="stat-value">{len(users)}</span>
                    </div>
                    <div class="stat-item">
                        <span class="stat-label">{'المستخدمون النشطون' if lang == 'ar' else 'Active Users'}</span>
                        <span class="stat-value">{sum(1 for u in users if u.is_active)}</span>
                    </div>
                    <div class="stat-item">
                        <span class="stat-label">{'المستخدمون غير النشطين' if lang == 'ar' else 'Inactive Users'}</span>
                        <span class="stat-value">{sum(1 for u in users if not u.is_active)}</span>
                    </div>
                </div>
            </div>
            
            <table>
                <thead>
                    <tr>
                        {''.join([f'<th>{h}</th>' for h in headers[lang]])}
                    </tr>
                </thead>
                <tbody>
                    {rows_html}
                </tbody>
            </table>
            <div class="footer">
                <p>{'تم التوليد بواسطة نظام Mcidia' if lang == 'ar' else 'Generated by Mcidia System'} | {timestamp}</p>
            </div>
        </body>
        </html>
        """
        
        # Convert HTML to PDF using weasyprint
        pdf_bytes = HTML(string=html_content).write_pdf()
        
        return send_file(
            BytesIO(pdf_bytes),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'users_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.pdf'
        )
    except Exception as e:
        print(f"PDF Export Error: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'خطأ في تصدير PDF / Error exporting PDF: {str(e)}', 'danger')
        return redirect(url_for('admin.users.index'))


@users_bp.route('/')
@login_required
@role_required('system_admin')
def index():
    """List all users with filtering"""
    db = get_db()
    lang = get_lang()
    
    # Get filter parameters
    role_filter = request.args.get('role')
    status_filter = request.args.get('status')
    plan_filter = request.args.get('plan')
    search = request.args.get('search', '')
    
    # Build query
    query = db.session.query(User)
    
    if role_filter:
        query = query.join(Role).filter(Role.name == role_filter)
    
    if status_filter:
        if status_filter == 'active':
            query = query.filter(User.is_active == True)
        elif status_filter == 'inactive':
            query = query.filter(User.is_active == False)
    
    if plan_filter:
        query = query.join(SubscriptionPlan).filter(SubscriptionPlan.name == plan_filter)
    
    if search:
        query = query.filter(
            (User.username.ilike(f'%{search}%')) |
            (User.email.ilike(f'%{search}%')) |
            (User.company_name.ilike(f'%{search}%'))
        )
    
    users = query.order_by(User.created_at.desc()).all()
    
    # Get all roles and plans for filters
    roles = db.session.query(Role).all()
    plans = db.session.query(SubscriptionPlan).all()
    
    return render_template(
        'admin/users/index.html',
        users=users,
        roles=roles,
        plans=plans,
        lang=lang,
        current_role_filter=role_filter,
        current_status_filter=status_filter,
        current_plan_filter=plan_filter,
        current_search=search
    )

@users_bp.route('/<int:user_id>')
@login_required
@role_required('system_admin')
def detail(user_id):
    """User detail page"""
    db = get_db()
    lang = get_lang()
    
    user = db.session.query(User).get_or_404(user_id)
    projects = db.session.query(Project).filter_by(user_id=user_id).order_by(Project.created_at.desc()).all()
    transactions = db.session.query(Transaction).filter_by(user_id=user_id).order_by(Transaction.created_at.desc()).all()
    ai_logs = db.session.query(AILog).filter_by(user_id=user_id).order_by(AILog.created_at.desc()).limit(20).all()
    
    roles = db.session.query(Role).all()
    plans = db.session.query(SubscriptionPlan).all()
    organizations = db.session.query(Organization).filter_by(is_active=True).all()
    
    return render_template(
        'admin/users/detail.html',
        user=user,
        projects=projects,
        transactions=transactions,
        ai_logs=ai_logs,
        roles=roles,
        plans=plans,
        organizations=organizations,
        lang=lang
    )

@users_bp.route('/<int:user_id>/update', methods=['POST'])
@login_required
@role_required('system_admin')
def update(user_id):
    """Update user details"""
    db = get_db()
    lang = get_lang()
    
    user = db.session.query(User).get_or_404(user_id)
    
    # Update fields
    user.username = request.form.get('username', user.username)
    user.email = request.form.get('email', user.email)
    user.phone = request.form.get('phone', user.phone)
    user.company_name = request.form.get('company_name', user.company_name)
    
    # Update role
    role_id = request.form.get('role_id')
    if role_id:
        user.role_id = int(role_id)
    
    # Update subscription plan
    plan_id = request.form.get('subscription_plan_id')
    if plan_id:
        user.subscription_plan_id = int(plan_id)
    
    # Update organization
    org_id = request.form.get('organization_id')
    if org_id:
        user.organization_id = int(org_id) if org_id != 'none' else None
    
    # Update active status
    user.is_active = request.form.get('is_active') == 'on'
    
    try:
        db.session.commit()
        flash('تم تحديث المستخدم بنجاح / User updated successfully' if lang == 'ar' else 'User updated successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'خطأ: {str(e)} / Error: {str(e)}' if lang == 'ar' else f'Error: {str(e)}', 'danger')
    
    return redirect(url_for('admin.users.detail', user_id=user_id))

@users_bp.route('/<int:user_id>/reset-password', methods=['POST'])
@login_required
@role_required('system_admin')
def reset_password(user_id):
    """Reset user password"""
    db = get_db()
    lang = get_lang()
    
    user = db.session.query(User).get_or_404(user_id)
    new_password = request.form.get('new_password')
    
    if new_password and len(new_password) >= 6:
        user.set_password(new_password)
        db.session.commit()
        flash('تم إعادة تعيين كلمة المرور بنجاح / Password reset successfully' if lang == 'ar' else 'Password reset successfully', 'success')
    else:
        flash('كلمة المرور يجب أن تكون 6 أحرف على الأقل / Password must be at least 6 characters' if lang == 'ar' else 'Password must be at least 6 characters', 'danger')
    
    return redirect(url_for('admin.users.detail', user_id=user_id))

@users_bp.route('/create', methods=['GET', 'POST'])
@login_required
@role_required('system_admin')
def create():
    """Create new user"""
    db = get_db()
    lang = get_lang()
    
    if request.method == 'GET':
        roles = db.session.query(Role).all()
        plans = db.session.query(SubscriptionPlan).all()
        organizations = db.session.query(Organization).filter_by(is_active=True).all()
        return render_template('admin/users/create.html', roles=roles, plans=plans, organizations=organizations, lang=lang)
    
    # POST - Create user
    try:
        org_id = int(request.form.get('organization_id')) if request.form.get('organization_id') and request.form.get('organization_id') != 'none' else None
        org_role = request.form.get('organization_role', 'member')
        
        new_user = User(
            username=request.form.get('username'),
            email=request.form.get('email'),
            phone=request.form.get('phone'),
            company_name=request.form.get('company_name'),
            role_id=int(request.form.get('role_id')),
            subscription_plan_id=int(request.form.get('subscription_plan_id')) if request.form.get('subscription_plan_id') else None,
            organization_id=org_id,
            is_active=request.form.get('is_active') == 'on'
        )
        new_user.set_password(request.form.get('password'))
        
        db.session.add(new_user)
        db.session.flush()  # Flush to get the user ID
        
        # If user is assigned to an organization, create organization membership
        if org_id:
            if org_role not in ['member', 'owner', 'consultant', 'admin']:
                org_role = 'member'  # Default to member if invalid role
            
            membership = OrganizationMembership(
                user_id=new_user.id,
                organization_id=org_id,
                membership_role=org_role,
                is_active=True,
                joined_at=datetime.utcnow()
            )
            db.session.add(membership)
        
        db.session.commit()
        
        flash('تم إنشاء المستخدم بنجاح / User created successfully' if lang == 'ar' else 'User created successfully', 'success')
        return redirect(url_for('admin.users.detail', user_id=new_user.id))
        
    except Exception as e:
        db.session.rollback()
        flash(f'خطأ: {str(e)} / Error: {str(e)}' if lang == 'ar' else f'Error: {str(e)}', 'danger')
        return redirect(url_for('admin.users.create'))

@users_bp.route('/<int:user_id>/delete', methods=['POST'])
@login_required
@role_required('system_admin')
def delete(user_id):
    """Delete user (soft delete)"""
    db = get_db()
    lang = get_lang()
    
    user = db.session.query(User).get_or_404(user_id)
    
    # Soft delete by deactivating
    user.is_active = False
    db.session.commit()
    
    flash('تم تعطيل المستخدم بنجاح / User deactivated successfully' if lang == 'ar' else 'User deactivated successfully', 'success')
    return redirect(url_for('admin.users.index'))
